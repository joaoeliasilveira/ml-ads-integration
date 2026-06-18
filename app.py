from flask import Flask, request, jsonify, render_template, redirect, session, abort
import requests
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, timezone, timedelta
import hashlib
import json as json_lib
import functools
from concurrent.futures import ThreadPoolExecutor, as_completed
import random

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "petina-dashboard-2026-secret")

# Compressão gzip automática nas respostas
from flask_compress import Compress
Compress(app)

try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass

CLIENT_ID     = os.environ.get("ML_CLIENT_ID")
CLIENT_SECRET = os.environ.get("ML_CLIENT_SECRET")
REDIRECT_URI  = os.environ.get("ML_REDIRECT_URI")
DATABASE_URL  = os.environ.get("DATABASE_URL")

TOKEN_TTL = timedelta(hours=5, minutes=30)

def get_db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sellers (
            user_id TEXT PRIMARY KEY,
            nickname TEXT,
            access_token TEXT,
            refresh_token TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # Tabela de usuários do dashboard
    cur.execute("""
        CREATE TABLE IF NOT EXISTS dashboard_users (
            username    TEXT PRIMARY KEY,
            password    TEXT NOT NULL,
            role        TEXT NOT NULL DEFAULT 'secondary',
            permissions JSONB DEFAULT '{}'::jsonb,
            created_at  TIMESTAMP DEFAULT NOW()
        )
    """)
    # Criar usuário master padrão se não existir
    cur.execute("SELECT COUNT(*) as cnt FROM dashboard_users WHERE role='master'")
    if cur.fetchone()['cnt'] == 0:
        import hashlib
        default_pwd = hashlib.sha256("petina2026".encode()).hexdigest()
        cur.execute(
            "INSERT INTO dashboard_users (username, password, role) VALUES (%s, %s, 'master') ON CONFLICT DO NOTHING",
            ("master", default_pwd)
        )
    # Tabela de cache para reduzir chamadas à API do ML
    cur.execute("""
        CREATE TABLE IF NOT EXISTS api_cache (
            cache_key  TEXT PRIMARY KEY,
            data       JSONB NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # Índice para limpeza eficiente por expires_at
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_api_cache_expires
        ON api_cache (expires_at)
    """)
    # Tabela de histórico diário de métricas por campanha
    cur.execute("""
        CREATE TABLE IF NOT EXISTS campaign_metrics_history (
            id            SERIAL PRIMARY KEY,
            seller_id     TEXT NOT NULL,
            campaign_id   TEXT NOT NULL,
            campaign_name TEXT,
            date          DATE NOT NULL,
            spend         NUMERIC DEFAULT 0,
            revenue       NUMERIC DEFAULT 0,
            roas          NUMERIC DEFAULT 0,
            acos          NUMERIC DEFAULT 0,
            clicks        INTEGER DEFAULT 0,
            impressions   INTEGER DEFAULT 0,
            orders        INTEGER DEFAULT 0,
            cvr           NUMERIC DEFAULT 0,
            recorded_at   TIMESTAMP DEFAULT NOW(),
            UNIQUE(seller_id, campaign_id, date)
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_camp_history_lookup
        ON campaign_metrics_history (seller_id, campaign_id, date DESC)
    """)
    conn.commit()
    cur.close()
    conn.close()

try:
    init_db()
except Exception as e:
    print("Erro ao inicializar banco: " + str(e))

# ─── AUTH HELPERS ─────────────────────────────────────────────────
def hash_pwd(pwd):
    return hashlib.sha256(pwd.encode()).hexdigest()

def get_current_user():
    return session.get("dash_user")

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("dash_user"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

def master_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        user = session.get("dash_user")
        if not user or user.get("role") != "master":
            if request.path.startswith("/api/"):
                return jsonify({"error": "forbidden"}), 403
            return redirect("/")
        return f(*args, **kwargs)
    return decorated

def can_see_seller(user, seller_id):
    if not user: return False
    if user.get("role") == "master": return True
    perms = user.get("permissions", {})
    sellers = perms.get("sellers", [])
    return not sellers or str(seller_id) in [str(s) for s in sellers]

def can_see_tab(user, tab):
    if not user: return False
    if user.get("role") == "master": return True
    perms = user.get("permissions", {})
    tabs  = perms.get("tabs", [])
    return not tabs or tab in tabs

# ─── CACHE HELPERS ────────────────────────────────────────────────
CACHE_TTL_SHORT   = 5   # dados voláteis (Hoje)
CACHE_TTL_MINUTES = 30  # dados históricos (padrão)
CACHE_TTL_LONG    = 60  # dados estáticos (produtos/reputação)

def cache_ttl(date_from, date_to):
    """Retorna TTL baseado no período: Hoje=5min, resto=30min."""
    today = datetime.now(timezone.utc).date().isoformat()
    if date_from == date_to == today:
        return CACHE_TTL_SHORT
    return CACHE_TTL_MINUTES

def cache_get(key):
    """Busca valor do cache. Retorna None se nao existir ou expirado."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "SELECT data FROM api_cache WHERE cache_key=%s AND expires_at > NOW()",
            (key,)
        )
        row = cur.fetchone()
        cur.close(); conn.close()
        return row["data"] if row else None
    except Exception:
        return None

def cache_set(key, data, ttl_minutes=CACHE_TTL_MINUTES):
    """Salva valor no cache com TTL."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            """INSERT INTO api_cache (cache_key, data, expires_at)
               VALUES (%s, %s, NOW() + INTERVAL '%s minutes')
               ON CONFLICT (cache_key) DO UPDATE
               SET data=%s, expires_at=NOW() + INTERVAL '%s minutes'""",
            (key, json_lib.dumps(data), ttl_minutes, json_lib.dumps(data), ttl_minutes)
        )
        if random.random() < 0.01:
            cur.execute('DELETE FROM api_cache WHERE expires_at < NOW()')
        conn.commit(); cur.close(); conn.close()
    except Exception:
        pass  # cache é best-effort

def cache_ttl(date_from, date_to):
    today = datetime.now(timezone.utc).date().isoformat()
    return 5 if date_to >= today else 30

def cache_key_sales(user_id, date_from, date_to):
    return f"sales:{user_id}:{date_from}:{date_to}"

def cache_key_ads(user_id, date_from, date_to):
    return f"ads:{user_id}:{date_from}:{date_to}"

def cache_key_products(user_id):
    return f"products:{user_id}"

# ─────────────────────────────────────────────────────────────────────
def cache_cleanup():
    """Remove entradas expiradas do cache. Chamar periodicamente."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM api_cache WHERE expires_at < NOW()")
        deleted = cur.rowcount
        conn.commit(); cur.close(); conn.close()
        return deleted
    except Exception:
        return 0

# Limpeza automática: disparada a cada N requests (probabilística)
import random
_cleanup_counter = 0
def maybe_cleanup_cache():
    global _cleanup_counter
    _cleanup_counter += 1
    if _cleanup_counter % 50 == 0:  # a cada ~50 requests
        cache_cleanup()

def check_seller_access(user_id):
    """Verifica se usuário logado tem acesso ao seller."""
    user = get_current_user()
    if not user:
        return False, (jsonify({"error": "unauthorized"}), 401)
    if not can_see_seller(user, user_id):
        return False, (jsonify({"error": "forbidden", "message": "Acesso nao autorizado a este seller"}), 403)
    return True, None

# ─── AUTH ROUTES ───────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    error = ""
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","").strip()
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute("SELECT * FROM dashboard_users WHERE username=%s", (username,))
            user = cur.fetchone()
            cur.close(); conn.close()
            if user and user["password"] == hash_pwd(password):
                session["dash_user"] = {
                    "username":    user["username"],
                    "role":        user["role"],
                    "permissions": user["permissions"] or {}
                }
                return redirect("/")
            error = "Usuário ou senha incorretos."
        except Exception as e:
            error = "Erro ao autenticar: " + str(e)
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ─── ADMIN ROUTES ──────────────────────────────────────────────────
@app.route("/admin")
@master_required
def admin():
    return render_template("admin.html")

@app.route("/api/admin/users", methods=["GET"])
@master_required
def admin_list_users():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT username, role, permissions, created_at FROM dashboard_users ORDER BY role, username")
    users = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(users)

@app.route("/api/admin/users", methods=["POST"])
@master_required
def admin_create_user():
    data = request.json or {}
    username = data.get("username","").strip()
    password = data.get("password","").strip()
    role     = data.get("role","secondary")
    perms    = data.get("permissions", {})
    if not username or not password:
        return jsonify({"error": "username e password obrigatorios"}), 400
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute(
            "INSERT INTO dashboard_users (username, password, role, permissions) VALUES (%s,%s,%s,%s)",
            (username, hash_pwd(password), role, json_lib.dumps(perms))
        )
        conn.commit(); cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/admin/users/<username>", methods=["PUT"])
@master_required
def admin_update_user(username):
    data    = request.json or {}
    updates = []
    params  = []
    if "permissions" in data:
        updates.append("permissions=%s")
        params.append(json_lib.dumps(data["permissions"]))
    if data.get("password"):
        updates.append("password=%s")
        params.append(hash_pwd(data["password"]))
    if not updates:
        return jsonify({"ok": True})
    params.append(username)
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"UPDATE dashboard_users SET {','.join(updates)} WHERE username=%s", params)
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<username>", methods=["DELETE"])
@master_required
def admin_delete_user(username):
    if username == "master":
        return jsonify({"error": "nao pode deletar o master"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM dashboard_users WHERE username=%s", (username,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/me")
@login_required
def admin_me():
    return jsonify(session.get("dash_user", {}))

# ─── HOME ──────────────────────────────────────────────────────────
@app.route("/")
@login_required
def home():
    return render_template("index.html")

@app.route("/auth")
def auth():
    auth_url = (
        "https://auth.mercadolivre.com.br/authorization"
        "?response_type=code"
        "&client_id=" + CLIENT_ID +
        "&redirect_uri=" + REDIRECT_URI +
        "&scope=read_ads+offline_access"
    )
    return redirect(auth_url)

@app.route("/callback")
def callback():
    code = request.args.get("code")
    if not code:
        return "Erro: codigo nao recebido", 400

    response = requests.post(
        "https://api.mercadolibre.com/oauth/token",
        data={
            "grant_type": "authorization_code",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "code": code,
            "redirect_uri": REDIRECT_URI
        }
    )

    token_data = response.json()
    access_token  = token_data.get("access_token")
    refresh_token = token_data.get("refresh_token")
    user_id       = str(token_data.get("user_id"))

    if not access_token:
        return "Erro ao obter token: " + str(token_data), 400

    user_resp = requests.get(
        "https://api.mercadolibre.com/users/" + user_id,
        headers={"Authorization": "Bearer " + access_token}
    )
    nickname = user_resp.json().get("nickname", "Seller " + user_id)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO sellers (user_id, nickname, access_token, refresh_token, updated_at)
        VALUES (%s, %s, %s, %s, NOW())
        ON CONFLICT (user_id) DO UPDATE
        SET access_token = EXCLUDED.access_token,
            refresh_token = EXCLUDED.refresh_token,
            nickname = EXCLUDED.nickname,
            updated_at = NOW()
    """, (user_id, nickname, access_token, refresh_token))
    conn.commit()
    cur.close()
    conn.close()

    return redirect("/?seller_added=" + nickname)


@app.route("/api/export/sales/<user_id>")
@login_required
def export_sales(user_id):
    """Exporta todos os pedidos do período em XLSX com todos os campos disponíveis."""
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to",   "")
    if not date_from or not date_to:
        today     = datetime.now(timezone.utc).date()
        date_to   = today.isoformat()
        date_from = (today - timedelta(days=30)).isoformat()

    # Buscar todos os pedidos
    def fetch_page_exp(offset, limit=50):
        r = requests.get(
            "https://api.mercadolibre.com/orders/search",
            params={
                "seller": user_id,
                "order.date_created.from": date_from + "T00:00:00.000-03:00",
                "order.date_created.to":   date_to   + "T23:59:59.000-03:00",
                "limit": limit, "offset": offset, "sort": "date_asc"
            },
            headers={"Authorization": "Bearer " + token},
            timeout=15
        )
        if not r.ok or not r.text: return [], 0
        d = r.json()
        return d.get("results", []), d.get("paging", {}).get("total", 0)

    first, total = fetch_page_exp(0)
    all_orders = list(first)
    if total > 50:
        offsets = list(range(50, min(total, 10000), 50))
        with ThreadPoolExecutor(max_workers=min(8, len(offsets))) as ex:
            futures = {ex.submit(fetch_page_exp, off): off for off in offsets}
            for f in as_completed(futures):
                res, _ = f.result()
                if res: all_orders.extend(res)
        # Deduplicar
        seen = set()
        unique = []
        for o in all_orders:
            oid = o.get("id")
            if oid and oid not in seen:
                seen.add(oid)
                unique.append(o)
        all_orders = unique

    # Ordenar por data
    all_orders.sort(key=lambda x: x.get("date_created", ""))

    # Gerar XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl nao instalado"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Cabeçalhos
    headers = [
        "N.° Pedido", "Pack ID", "Data Criação", "Data Fechamento",
        "Status", "Status Detail", "Substatus",
        "Item ID", "Item Título", "SKU", "Qtd", "Preço Unit.",
        "Total Amount", "Paid Amount", "Currency",
        "Forma Pagamento", "Parcelas", "Total Pago (payments)",
        "Frete Tipo", "Frete Status", "Data Envio", "Data Entrega",
        "Comprador ID", "Comprador Nickname",
        "Tags", "Feedback Comprador", "Feedback Vendedor",
        "Mediação", "Data Mediação",
        "Campaing", "Venda por ADS",
    ]

    # Estilo do cabeçalho
    header_fill = PatternFill("solid", fgColor="0d4a47")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Preencher dados - uma linha por item do pedido
    row_num = 2
    for o in all_orders:
        oid          = o.get("id", "")
        pack_id      = o.get("pack_id", "")
        date_created = (o.get("date_created") or "")[:19].replace("T", " ")
        date_closed  = (o.get("date_closed")  or "")[:19].replace("T", " ")
        status       = o.get("status", "")
        status_detail= o.get("status_detail", "")
        substatus    = o.get("substatus", "")
        total_amount = o.get("total_amount", 0) or 0
        paid_amount  = o.get("paid_amount",  0) or 0
        currency     = o.get("currency_id", "BRL")
        tags         = ", ".join(o.get("tags") or [])
        buyer        = o.get("buyer") or {}
        buyer_id     = buyer.get("id", "")
        buyer_nick   = buyer.get("nickname", "")

        # Pagamentos
        payments = o.get("payments") or []
        pay_type     = ", ".join(set(p.get("payment_type", "") for p in payments if p.get("payment_type")))
        pay_install  = max((p.get("installments", 1) or 1 for p in payments), default=1)
        pay_total    = sum(p.get("total_paid_amount", 0) or 0 for p in payments)

        # Envio
        shipping     = o.get("shipping") or {}
        ship_type    = shipping.get("shipping_option", {}).get("name", "") if isinstance(shipping.get("shipping_option"), dict) else ""
        ship_status  = shipping.get("status", "")
        ship_date    = (shipping.get("date_shipped") or "")[:19].replace("T", " ")
        deliver_date = (shipping.get("date_delivered") or "")[:19].replace("T", " ")

        # Feedback
        feedback     = o.get("feedback") or {}
        fb_buyer     = (feedback.get("purchase") or {}).get("rating", "")
        fb_seller    = (feedback.get("sale")     or {}).get("rating", "")

        # Mediação
        med          = o.get("mediations") or []
        med_flag     = "Sim" if med else "Não"
        med_date     = (med[0].get("date_started", "") if med else "")[:10]

        # Contexto (ADS)
        context      = o.get("context") or {}
        campaign     = context.get("campaign_id", "")
        is_ads       = "Sim" if context.get("ads_type") else "Não"

        items = o.get("order_items") or []
        if not items:
            items = [{}]

        for item in items:
            it        = item.get("item") or {}
            item_id   = it.get("id", "")
            item_title= it.get("title", "")
            item_sku  = item.get("item", {}).get("seller_custom_field", "") if item.get("item") else ""
            qty       = item.get("quantity", 1) or 1
            unit_price= item.get("unit_price", 0) or 0

            ws.append([
                oid, pack_id, date_created, date_closed,
                status, status_detail, substatus,
                item_id, item_title, item_sku, qty, unit_price,
                total_amount, paid_amount, currency,
                pay_type, pay_install, pay_total,
                ship_type, ship_status, ship_date, deliver_date,
                buyer_id, buyer_nick,
                tags, fb_buyer, fb_seller,
                med_flag, med_date,
                campaign, is_ads,
            ])
            row_num += 1
            # Após primeiro item: limpar campos de nível de order (evitar duplicação)
            total_amount = paid_amount = pay_total = ""
            pack_id = date_created = date_closed = ""
            status = status_detail = substatus = ""
            buyer_id = buyer_nick = tags = ""
            pay_type = pay_install = ""
            ship_type = ship_status = ship_date = deliver_date = ""
            fb_buyer = fb_seller = med_flag = med_date = ""
            campaign = is_ads = ""

    # Ajustar largura das colunas
    col_widths = [15,12,20,20,15,15,15,15,45,15,6,12,14,14,8,15,8,14,18,14,18,18,12,20,25,12,12,8,12,15,8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    # Salvar e retornar
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    filename = f"pedidos_{seller['nickname']}_{date_from}_{date_to}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route("/api/sellers")
@login_required
def get_sellers():
    user = get_current_user()
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT user_id, nickname FROM sellers ORDER BY nickname")
    all_s = [dict(s) for s in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify([s for s in all_s if can_see_seller(user, s["user_id"])])

@app.route("/api/sellers/<user_id>", methods=["DELETE"])
def delete_seller(user_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM sellers WHERE user_id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"success": True})

def refresh_token_if_needed(user_id, access_token, refresh_token, updated_at):
    now = datetime.now(timezone.utc)
    if updated_at.tzinfo is None:
        updated_at = updated_at.replace(tzinfo=timezone.utc)
    if now - updated_at < TOKEN_TTL:
        return access_token
    resp = requests.post(
        "https://api.mercadolibre.com/oauth/token",
        data={
            "grant_type": "refresh_token",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "refresh_token": refresh_token
        }
    )
    new_tokens = resp.json()
    new_access  = new_tokens.get("access_token", access_token)
    new_refresh = new_tokens.get("refresh_token", refresh_token)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE sellers SET access_token=%s, refresh_token=%s, updated_at=NOW()
        WHERE user_id=%s
    """, (new_access, new_refresh, user_id))
    conn.commit()
    cur.close()
    conn.close()
    return new_access

def get_seller_token(user_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sellers WHERE user_id = %s", (user_id,))
    seller = cur.fetchone()
    cur.close()
    conn.close()
    if not seller:
        return None, None
    token = refresh_token_if_needed(user_id, seller["access_token"], seller["refresh_token"], seller["updated_at"])
    return token, seller

def get_advertiser_id(user_id, token):
    resp = requests.get(
        "https://api.mercadolibre.com/advertising/advertisers",
        params={"product_id": "PADS"},
        headers={"Authorization": "Bearer " + token, "Api-Version": "1"}
    )
    if not resp.ok:
        return None
    data = resp.json()
    advertisers = data if isinstance(data, list) else data.get("advertisers", data.get("results", []))
    if advertisers and len(advertisers) > 0:
        return str(advertisers[0].get("id", advertisers[0].get("advertiser_id", "")))
    return None

def get_campaigns(user_id, token):
    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    endpoints = [
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns",
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/campaigns",
    ]
    for url in endpoints:
        resp = requests.get(url, headers={"Authorization": "Bearer " + token, "Api-Version": "1"})
        try:
            data = resp.json()
        except Exception:
            continue
        if not resp.ok:
            print("[ADS][ERRO] GET " + url + " HTTP " + str(resp.status_code))
            continue
        if isinstance(data, list):
            return data, url, aid
        if isinstance(data, dict) and "results" in data:
            return data["results"], url, aid
    return [], None, aid

def get_campaign_metrics(advertiser_id, camp_id, token, date_from, date_to, base_url):
    url = "https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/" + str(camp_id)
    resp = requests.get(url, params={
        "date_from": date_from,
        "date_to": date_to,
        "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"
    }, headers={"Authorization": "Bearer " + token, "Api-Version": "2"})
    if not resp.ok:
        print("[METRICS][ERRO] " + url + " HTTP " + str(resp.status_code) + " " + resp.text[:200])
        return {}
    try:
        return resp.json()
    except Exception:
        return {}

@app.route("/api/ads/<user_id>")
def get_ads(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "2026-05-01")
    date_to   = request.args.get("date_to",   "2026-05-26")
    campaigns, base_url, aid = get_campaigns(user_id, token)

    result = []
    total_spend = total_revenue = total_clicks = total_impressions = 0

    def fetch_camp_metrics(camp):
        camp_id = camp.get("id")
        metrics = get_campaign_metrics(aid, camp_id, token, date_from, date_to, base_url or "")

        # Formato: response tem objeto "metrics" dentro
        if isinstance(metrics, dict) and "metrics" in metrics:
            m = metrics["metrics"]
        elif isinstance(metrics, list) and len(metrics) > 0:
            m = metrics[0].get("metrics", metrics[0])
        else:
            m = metrics if isinstance(metrics, dict) else {}

        return camp, camp_id, metrics, m

    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(fetch_camp_metrics, camp): camp for camp in campaigns}
        camps_results = []
        for future in as_completed(futures):
            try:
                camps_results.append(future.result())
            except Exception as e:
                print("[ADS][ERRO] campanha:", e)

    for camp, camp_id, metrics, m in camps_results:

        spend   = m.get("cost", 0)
        revenue = m.get("total_amount", m.get("direct_amount", m.get("revenue", 0)))
        clicks  = m.get("clicks", 0)
        imps    = m.get("prints", m.get("impressions", 0))
        orders  = m.get("advertising_items_quantity", m.get("direct_items_quantity", 0))
        cvr     = m.get("cvr", 0)

        roas = round(revenue / spend, 2) if spend > 0 else 0
        acos = round((spend / revenue) * 100, 1) if revenue > 0 else 0
        total_spend += spend
        total_revenue += revenue
        total_clicks += clicks
        total_impressions += imps
        # roas_target vem do objeto da campanha (não das métricas)
        roas_target = camp.get("roas_target", None)
        if roas_target is None and isinstance(metrics, dict):
            roas_target = metrics.get("roas_target", None)
        if roas_target is not None:
            try:
                roas_target = round(float(roas_target), 2)
            except Exception:
                roas_target = None

        result.append({
            "id": camp_id,
            "name": camp.get("name", "Campanha " + str(camp_id)),
            "status": camp.get("status", "unknown"),
            "spend": round(spend, 2),
            "revenue": round(revenue, 2),
            "clicks": clicks,
            "impressions": imps,
            "roas": roas,
            "acos": acos,
            "roas_target": roas_target,
            "orders": orders,
            "cvr": round(float(cvr), 2) if cvr else 0,
        })

    return jsonify({
        "seller_id": user_id,
        "nickname": seller["nickname"],
        "summary": {
            "spend": round(total_spend, 2),
            "revenue": round(total_revenue, 2),
            "clicks": total_clicks,
            "impressions": total_impressions,
            "roas": round(total_revenue / total_spend, 2) if total_spend > 0 else 0,
            "acos": round((total_spend / total_revenue) * 100, 1) if total_revenue > 0 else 0
        },
        "campaigns": result
    })

@app.route("/api/ads/<user_id>/daily")
def get_ads_daily(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    _df = request.args.get("date_from",""); _dt = request.args.get("date_to","")
    _ck = f"ads_daily:{user_id}:{_df}:{_dt}"
    _cached = cache_get(_ck)
    if _cached: return jsonify(_cached)
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "2026-05-01")
    date_to   = request.args.get("date_to",   "2026-05-26")
    campaigns, base_url, aid = get_campaigns(user_id, token)
    daily_map = {}

    for camp in campaigns:
        camp_id = camp.get("id")
        metrics = get_campaign_metrics(aid, camp_id, token, date_from, date_to, base_url or "")
        if isinstance(metrics, list):
            for day in metrics:
                date = day.get("date", "")
                if not date:
                    continue
                if date not in daily_map:
                    daily_map[date] = {"date": date, "cost": 0, "revenue": 0, "clicks": 0, "impressions": 0}
                daily_map[date]["cost"]        += day.get("cost", 0)
                daily_map[date]["revenue"]     += day.get("revenue", 0)
                daily_map[date]["clicks"]      += day.get("clicks", 0)
                daily_map[date]["impressions"] += day.get("impressions", 0)

    days_list = sorted(daily_map.values(), key=lambda x: x["date"])
    for d in days_list:
        d["cost"]    = round(d["cost"], 2)
        d["revenue"] = round(d["revenue"], 2)

    _res_d = {
        "seller_id": user_id,
        "nickname": seller["nickname"],
        "date_from": date_from,
        "date_to": date_to,
        "days": days_list
    }
    try: cache_set(_ck, _res_d, cache_ttl(_df, _dt))
    except Exception: pass
    return jsonify(_res_d)

@app.route("/api/reputation/<user_id>")
def get_reputation(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    rep_resp = requests.get(
        "https://api.mercadolibre.com/users/" + user_id,
        headers={"Authorization": "Bearer " + token}
    )
    rep_data = rep_resp.json() if rep_resp.ok else {}
    reputation = rep_data.get("seller_reputation", {})
    metrics_rep = reputation.get("metrics", {})
    transactions = reputation.get("transactions", {})

    return jsonify({
        "seller_id": user_id,
        "nickname": seller["nickname"],
        "reputation": {
            "level": reputation.get("level_id", ""),
            "power_seller": rep_data.get("power_seller_status", ""),
            "cancellations": metrics_rep.get("cancellations", {}).get("rate", 0),
            "claims": metrics_rep.get("claims", {}).get("rate", 0),
            "delayed": metrics_rep.get("delayed_handling_time", {}).get("rate", 0),
            "sales_completed": transactions.get("completed", 0),
            "ratings": reputation.get("ratings", {})
        }
    })

@app.route("/api/sales/<user_id>")
@login_required
def get_sales(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to",   "")
    if not date_from or not date_to:
        today     = datetime.now(timezone.utc).date()
        date_to   = today.isoformat()
        date_from = (today - timedelta(days=30)).isoformat()

    # Cache
    ck = cache_key_sales(user_id, date_from, date_to)
    cached = cache_get(ck)
    if cached:
        return jsonify(cached)

    from datetime import date as ddate, timedelta as tdelta

    # ── BUSCA DE PEDIDOS ────────────────────────────────────────────────────────
    # Usa date_created (igual ao ML): captura todos os pedidos feitos no período
    # incluindo os ainda em trânsito (que não têm date_closed)
    # NÃO filtra por status: ML conta todos exceto invalid/cancelled para GMV

    def fetch_page(dfrom, dto, offset, limit=50, date_field="date_created"):
        r = requests.get(
            "https://api.mercadolibre.com/orders/search",
            params={
                "seller": user_id,
                f"order.{date_field}.from": dfrom + "T00:00:00.000-03:00",
                f"order.{date_field}.to":   dto   + "T23:59:59.000-03:00",
                "limit": limit, "offset": offset, "sort": "date_asc"
            },
            headers={"Authorization": "Bearer " + token},
            timeout=15
        )
        if not r.ok or not r.text: return [], 0
        d = r.json()
        return d.get("results", []), d.get("paging", {}).get("total", 0)

    def fetch_all(dfrom, dto, date_field="date_created"):
        first, total = fetch_page(dfrom, dto, 0, date_field=date_field)
        if not first: return []
        all_orders = list(first)
        if total > 50:
            offsets = list(range(50, min(total, 10000), 50))
            with ThreadPoolExecutor(max_workers=min(8, len(offsets))) as ex:
                futures = {ex.submit(fetch_page, dfrom, dto, off, date_field=date_field): off
                           for off in offsets}
                failed = []
                for f in as_completed(futures):
                    res, _ = f.result()
                    if res: all_orders.extend(res)
                    else:   failed.append(futures[f])
            for off in failed:
                res, _ = fetch_page(dfrom, dto, off, date_field=date_field)
                if res: all_orders.extend(res)
        seen = set(); unique = []
        for o in all_orders:
            oid = o.get("id")
            if oid and oid not in seen:
                seen.add(oid); unique.append(o)
        return unique

    def fetch_all_merged(dfrom, dto):
        """Busca por date_closed (primario, ML AI confirmou) + date_created (secundario).
        Une e deduplica por order_id.
        """
        with ThreadPoolExecutor(max_workers=2) as ex:
            f1 = ex.submit(fetch_all, dfrom, dto, "date_closed")
            f2 = ex.submit(fetch_all, dfrom, dto, "date_created")
            by_closed  = f1.result()
            by_created = f2.result()
        seen = set(); merged = []
        for o in by_closed + by_created:
            oid = o.get("id")
            if oid and oid not in seen:
                seen.add(oid); merged.append(o)
        return merged

    def order_gmv(o):
        """GMV = total_amount.
        Nota: acrescimo de parcelamento nao tem campo direto confiavel no /orders/search.
        Diferenca residual de ~R$247 vs ML e o acrescimo de parcelamento (Col J planilha).
        """
        val = o.get("total_amount") or 0
        if val > 0: return val
        items = o.get("order_items") or []
        if items:
            total = sum((i.get("gross_price") or i.get("unit_price") or 0) *
                        (i.get("quantity") or 1) for i in items)
            if total > 0: return total
        return 0

    def count_unique(order_list):
        """Conta orders individualmente — o ML nao agrupa packs para Qtd de Vendas."""
        return len(order_list)

    def sum_units(order_list):
        total = 0
        for o in order_list:
            for item in (o.get("order_items") or []):
                total += item.get("quantity") or 1
        return total

    # Buscar período atual e anterior + visitas em paralelo
    d_from   = ddate.fromisoformat(date_from)
    d_to     = ddate.fromisoformat(date_to)
    delta    = (d_to - d_from).days + 1
    prev_from = (d_from - tdelta(days=delta)).isoformat()
    prev_to   = (d_from - tdelta(days=1)).isoformat()

    def _extract_visits(d):
        """Extrai total_visits de qualquer estrutura de resposta do ML.
        A API pode retornar total_visits na raiz ou dentro de d['results'].
        """
        # Raiz direta
        v = d.get("total_visits") or d.get("visits") or 0
        if v > 0:
            return v
        # Aninhado em results (estrutura observada no endpoint date_range)
        inner = d.get("results")
        if isinstance(inner, dict):
            v = inner.get("total_visits") or inner.get("visits") or 0
            if v > 0:
                return v
            # Fallback: somar visits_detail se existir
            detail = inner.get("visits_detail") or []
            if detail:
                total = sum(entry.get("total", 0) for entry in detail)
                if total > 0:
                    return total
        return 0

    def _get_visits():
        # Tentativa 1: endpoint time_window (ML recomendado para apps de terceiros)
        try:
            r = requests.get(
                f"https://api.mercadolibre.com/users/{user_id}/items_visits/time_window",
                params={"last": delta, "unit": "day", "ending": date_to},
                headers={"Authorization": "Bearer " + token}, timeout=8
            )
            if r.ok and r.text:
                v = _extract_visits(r.json())
                if v > 0: return v
        except: pass

        # Tentativa 2: endpoint com date_from/date_to explícitos (confirmado funcionando)
        try:
            r = requests.get(
                f"https://api.mercadolibre.com/users/{user_id}/items_visits",
                params={"date_from": date_from, "date_to": date_to},
                headers={"Authorization": "Bearer " + token}, timeout=8
            )
            if r.ok and r.text:
                v = _extract_visits(r.json())
                if v > 0: return v
        except: pass

        # Tentativa 3: com timezone explícito
        try:
            r = requests.get(
                f"https://api.mercadolibre.com/users/{user_id}/items_visits",
                params={
                    "date_from": date_from + "T00:00:00.000-03:00",
                    "date_to":   date_to   + "T23:59:59.000-03:00"
                },
                headers={"Authorization": "Bearer " + token}, timeout=8
            )
            if r.ok and r.text:
                v = _extract_visits(r.json())
                if v > 0: return v
        except: pass

        return 0

    with ThreadPoolExecutor(max_workers=3) as ex:
        # fetch_all_merged: combina date_created + date_closed para capturar
        # todos os pedidos incluindo os entregues criados antes do período
        f_cur  = ex.submit(fetch_all_merged, date_from, date_to)
        f_prev = ex.submit(fetch_all_merged, prev_from, prev_to)
        f_vis  = ex.submit(_get_visits)
        orders      = f_cur.result()
        prev_orders = f_prev.result()
        total_visits = f_vis.result()

    # ── MÉTRICAS PRINCIPAIS ─────────────────────────────────────────────────────
    gmv        = round(sum(order_gmv(o) for o in orders), 2)
    qtd_vendas = count_unique(orders)
    units      = sum_units(orders)
    avg_unit   = round(gmv / units, 2) if units > 0 else 0
    avg_sale   = round(gmv / qtd_vendas, 2) if qtd_vendas > 0 else 0
    conversion = round((qtd_vendas / total_visits) * 100, 2) if total_visits > 0 else 0

    # Cancelamentos vs Devolucoes
    # shipping.status nao vem preenchido no /orders/search para pedidos cancelados.
    # Solucao: para cada cancelado com shipping_id, buscar /shipments/{id} em paralelo
    # para checar se foi entregue antes de ser cancelado (= devolucao).

    CANCEL_STATUSES = {"cancelled"}
    all_cancelled = [o for o in orders if o.get("status") in CANCEL_STATUSES]

    def get_shipment_status(ship_id):
        try:
            r = requests.get(
                f"https://api.mercadolibre.com/shipments/{ship_id}",
                headers={"Authorization": "Bearer " + token},
                timeout=8
            )
            if r.ok and r.text:
                d = r.json()
                return d.get("status", ""), d.get("substatus", "")
        except Exception:
            pass
        return "", ""

    # Critério ML (confirmado pela IA):
    # DEVOLUÇÃO  = produto foi ENTREGUE ao comprador e depois entrou em fluxo de retorno
    # CANCELAMENTO = reembolso ANTES da entrega (not_delivered, lost, shipped, etc.)
    # Apenas "delivered" como status do shipment garante que houve entrega real.
    # "returned"/"returning" sozinhos podem ocorrer sem entrega prévia confirmada,
    # mas quando o shipment chegou a "delivered" antes = devolução.

    RETURN_STATUSES = {"delivered"}          # entregue = possível devolução
    CANCEL_STATUSES_SHIP = {               # nunca chegou = cancelamento
        "not_shipped", "pending",
        "shipped",                          # a caminho, não entregue
        "not_delivered",                    # tentativa falhou, devolvido ao remetente
        "lost",                             # perdido em trânsito
        "cancelled",
    }

    order_is_return = {}

    ship_ids = {}
    for o in all_cancelled:
        shipping = o.get("shipping") or {}
        sid = shipping.get("id")
        if sid:
            ship_ids[str(o.get("id", ""))] = str(sid)

    if ship_ids:
        with ThreadPoolExecutor(max_workers=min(16, len(ship_ids))) as ex:
            futures = {ex.submit(get_shipment_status, sid): oid
                       for oid, sid in ship_ids.items()}
            for f in as_completed(futures):
                oid = futures[f]
                s_status, s_substatus = f.result()
                # Devolução = foi entregue (delivered) ou está em processo de retorno
                # após ter sido entregue (returned/returning com substatus de retorno)
                is_ret = s_status in RETURN_STATUSES
                if not is_ret and s_status in {"returned", "returning", "return_to_sender"}:
                    # Só conta como devolução se o substatus indica retorno pós-entrega
                    is_ret = s_substatus in {"returning_to_sender", "delivered_to_sender",
                                             "return_success", "returning"}
                order_is_return[oid] = is_ret

    def is_return(o):
        """Devolucao = pedido que foi entregue antes de ser cancelado.
        Usa resultado do lookup de /shipments/{id} feito em paralelo acima.
        Fallback para cancel_detail se shipment nao disponivel.
        """
        oid = str(o.get("id", ""))
        # Resultado do lookup de shipment
        if oid in order_is_return:
            return order_is_return[oid]
        # Fallback: cancel_detail
        cd = o.get("cancel_detail") or {}
        cd_code = str(cd.get("code") or "").lower()
        cd_desc = str(cd.get("description") or "").lower()
        if cd_code in {"return", "not_received", "item_not_as_described",
                       "buyer_remorse", "damaged_item", "wrong_item"}:
            return True
        if "devol" in cd_desc or "return" in cd_desc:
            return True
        return False

    returns    = [o for o in all_cancelled if is_return(o)]
    cancelled  = [o for o in all_cancelled if not is_return(o)]

    qtd_cancel    = count_unique(cancelled)
    valor_cancel  = round(sum(order_gmv(o) for o in cancelled), 2)
    qtd_returns   = count_unique(returns)
    valor_returns = round(sum(order_gmv(o) for o in returns), 2)

    # ── PERÍODO ANTERIOR ────────────────────────────────────────────────────────
    prev_gmv   = round(sum(order_gmv(o) for o in prev_orders), 2)
    prev_qtd   = count_unique(prev_orders)
    prev_units = sum_units(prev_orders)
    prev_avg   = round(prev_gmv / prev_units, 2) if prev_units > 0 else 0

    def var(curr, prev):
        if prev and prev != 0:
            return round((curr - prev) / abs(prev) * 100, 1)
        return 0

    # ── VENDAS DIÁRIAS ──────────────────────────────────────────────────────────
    daily_map = {}
    for o in orders:
        day = (o.get("date_created") or "")[:10]
        if not day: continue
        if day not in daily_map:
            daily_map[day] = {"date": day, "gmv": 0, "qtd": 0, "units": 0}
        daily_map[day]["gmv"]   += order_gmv(o)
        daily_map[day]["qtd"]   += 1
        daily_map[day]["units"] += sum((i.get("quantity") or 1) for i in (o.get("order_items") or []))

    daily_sales = sorted(daily_map.values(), key=lambda x: x["date"])
    for d in daily_sales:
        d["gmv"] = round(d["gmv"], 2)

    # ── PRODUTOS MAIS VENDIDOS (top 50) ─────────────────────────────────────────
    ck_prod = f"products_base:{user_id}"
    products_map = {}
    try:
        all_item_ids = []
        scroll_id = None
        for _ in range(20):
            params = {"limit": 100}
            if scroll_id: params["scroll_id"] = scroll_id
            r = requests.get(
                f"https://api.mercadolibre.com/users/{user_id}/items/search",
                params=params,
                headers={"Authorization": "Bearer " + token}, timeout=8
            )
            if not r.ok: break
            d = r.json()
            batch = d.get("results", [])
            if not batch: break
            all_item_ids.extend(batch)
            scroll_id = d.get("scroll_id")
            if not scroll_id or len(batch) < 100: break

        for chunk in [all_item_ids[i:i+20] for i in range(0, len(all_item_ids), 20)]:
            r2 = requests.get(
                "https://api.mercadolibre.com/items",
                params={"ids": ",".join(chunk),
                        "attributes": "id,title,price,status,available_quantity,shipping,fulfillment,catalog_product_id,logistic_type"},
                headers={"Authorization": "Bearer " + token}, timeout=8
            )
            if not r2.ok: continue
            for entry in r2.json():
                body = entry.get("body", {})
                iid  = str(body.get("id", ""))
                if not iid: continue
                ff   = body.get("fulfillment") or {}
                sh   = body.get("shipping") or {}
                # logistic_type pode vir na raiz, dentro de shipping, ou dentro de fulfillment
                lt   = (body.get("logistic_type") or
                        sh.get("logistic_type") or
                        ff.get("logistic_type") or "")
                full_keywords = ("fulfillment", "meli_fulfillment")
                is_full = bool(
                    lt in full_keywords or
                    ff.get("fulfillment_id") or
                    ff.get("status") == "active" or
                    sh.get("fulfillment") or
                    any(k in str(sh).lower() for k in full_keywords) or
                    any(k in str(ff).lower() for k in full_keywords)
                )
                print(f"[FULL_DEBUG] {iid} lt={lt!r} ff={ff} sh_lt={sh.get('logistic_type')!r} is_full={is_full}")
                products_map[iid] = {
                    "item_id": iid,
                    "title":   body.get("title", ""),
                    "price":   body.get("price", 0),
                    "status":  body.get("status", ""),
                    "stock_total": body.get("available_quantity", 0),
                    "is_full": is_full,
                    "catalog_product_id": body.get("catalog_product_id", ""),
                    "qty": 0, "revenue": 0, "units": 0,
                    "ads_qty": 0, "organic_qty": 0
                }
    except: pass

    for o in orders:
        for item in (o.get("order_items") or []):
            iid = str(item.get("item", {}).get("id", "") or "")
            if not iid: continue
            if iid not in products_map:
                products_map[iid] = {
                    "item_id": iid, "title": item.get("item", {}).get("title", ""),
                    "price": item.get("unit_price", 0), "status": "active",
                    "stock_total": 0, "is_full": False, "catalog_product_id": "",
                    "qty": 0, "revenue": 0, "units": 0, "ads_qty": 0, "organic_qty": 0
                }
            qty = item.get("quantity") or 1
            rev = (item.get("unit_price") or 0) * qty
            products_map[iid]["qty"]     += 1
            products_map[iid]["units"]   += qty
            products_map[iid]["revenue"] += rev

    top_products = sorted(products_map.values(), key=lambda x: x["revenue"], reverse=True)[:50]

    # ── RESULTADO FINAL ─────────────────────────────────────────────────────────
    result = {
        "seller_id": user_id,
        "nickname":  seller["nickname"],
        "period":    {"date_from": date_from, "date_to": date_to},
        "summary": {
            "vendas_brutas":       gmv,
            "unidades_vendidas":   units,
            "preco_medio_unidade": avg_unit,
            "qtd_vendas":          qtd_vendas,
            "preco_medio_venda":   avg_sale,
            "total_visits":        total_visits,
            "conversion":          conversion,
            "qtd_canceladas":      qtd_cancel,
            "valor_canceladas":    valor_cancel,
            "qtd_devolvidas":      qtd_returns,
            "valor_devolvidas":    valor_returns,
        },
        "comparison": {
            "prev_period":        {"date_from": prev_from, "date_to": prev_to},
            "prev_vendas_brutas": prev_gmv,
            "prev_qtd_vendas":    prev_qtd,
            "prev_unidades":      prev_units,
            "prev_avg_unit":      prev_avg,
            "var_vendas_brutas":  var(gmv, prev_gmv),
            "var_qtd_vendas":     var(qtd_vendas, prev_qtd),
            "var_unidades":       var(units, prev_units),
            "var_avg_unit":       var(avg_unit, prev_avg),
            "var_conversion":     var(conversion, 0),
        },
        "daily_sales":  daily_sales,
        "top_products": top_products
    }

    maybe_cleanup_cache()
    try: cache_set(ck, result, cache_ttl(date_from, date_to))
    except Exception: pass
    return jsonify(result)

@app.route("/api/promotions/<user_id>")
def get_promotions(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    _ck_p = f"promotions:{user_id}"
    _cached_p = cache_get(_ck_p)
    if _cached_p: return jsonify(_cached_p)
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    promotions = []
    status_info = {}

    # Endpoint correto conforme documentacao ML: app_version=v2
    r1 = requests.get(
        "https://api.mercadolibre.com/seller-promotions/users/" + user_id,
        params={"app_version": "v2"},
        headers={"Authorization": "Bearer " + token}
    )
    status_info["seller_promotions_v2"] = r1.status_code
    if r1.ok and r1.text:
        try:
            d = r1.json()
            raw = d if isinstance(d, list) else d.get("results", d.get("promotions", []))
            # Filtra apenas promos ativas
            promotions = [p for p in raw if p.get("status") in ("started", "active", "candidate")]
            if not promotions:
                promotions = raw  # mostra todas se nenhuma ativa
        except Exception:
            pass

    # Fallback: sem app_version
    if not promotions:
        r2 = requests.get(
            "https://api.mercadolibre.com/seller-promotions/users/" + user_id + "/promotions",
            params={"app_version": "v2"},
            headers={"Authorization": "Bearer " + token}
        )
        status_info["seller_promotions_path_v2"] = r2.status_code
        if r2.ok and r2.text:
            try:
                d = r2.json()
                promotions = d if isinstance(d, list) else d.get("results", d.get("promotions", []))
            except Exception:
                pass

    result = []
    for p in promotions[:20]:
        result.append({
            "id": str(p.get("id", p.get("deal_id", ""))),
            "name": p.get("name", p.get("description", p.get("deal_print_id", "Promocao"))),
            "type": p.get("type", p.get("deal_type", p.get("promotion_type", ""))),
            "status": p.get("status", ""),
            "discount": p.get("value", p.get("discount_meli_amount", p.get("percent_off", 0))),
            "start_date": str(p.get("start_date", p.get("from_date", p.get("from", ""))))[:10],
            "end_date": str(p.get("finish_date", p.get("to_date", p.get("to", ""))))[:10],
            "items_count": p.get("items_count", p.get("affected_items", 0))
        })

    _res_promo = {
        "seller_id": user_id,
        "nickname": seller["nickname"],
        "total": len(result),
        "promotions": result,
        "status_info": status_info
    }
    try: cache_set(_ck_p, _res_promo, 10)
    except Exception: pass
    try: cache_set(f"promotions:{user_id}", _res_promo, CACHE_TTL_MINUTES)
    except Exception: pass
    return jsonify(_res_promo)

@app.route("/api/metrics/<user_id>")
def get_metrics(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "2026-05-01")
    date_to   = request.args.get("date_to",   "2026-05-26")

    rep_resp = requests.get(
        "https://api.mercadolibre.com/users/" + user_id,
        headers={"Authorization": "Bearer " + token}
    )
    rep_data = rep_resp.json() if rep_resp.ok else {}
    reputation = rep_data.get("seller_reputation", {})
    metrics_rep = reputation.get("metrics", {})
    transactions = reputation.get("transactions", {})

    sales_resp = requests.get(
        "https://api.mercadolibre.com/orders/search",
        params={
            "seller": user_id,
            "order.date_created.from": date_from + "T00:00:00.000-03:00",
            "order.date_created.to":   date_to + "T23:59:59.000-03:00",
            "order.status": "paid",
            "limit": 50
        },
        headers={"Authorization": "Bearer " + token}
    )
    sales_data   = sales_resp.json() if sales_resp.ok else {}
    orders       = sales_data.get("results", [])
    total_orders = sales_data.get("paging", {}).get("total", 0)
    gmv          = sum(o.get("total_amount", 0) for o in orders)

    visits_resp = requests.get(
        "https://api.mercadolibre.com/users/" + user_id + "/items_visits",
        params={"date_from": date_from, "date_to": date_to},
        headers={"Authorization": "Bearer " + token}
    )
    visits_data  = visits_resp.json() if visits_resp.ok else {}
    total_visits = visits_data.get("total_visits", 0)

    return jsonify({
        "seller_id": user_id,
        "nickname": seller["nickname"],
        "period": {"date_from": date_from, "date_to": date_to},
        "sales": {
            "total_orders": total_orders,
            "gmv": round(gmv, 2),
            "avg_ticket": round(gmv / total_orders, 2) if total_orders > 0 else 0
        },
        "visits": {
            "total": total_visits,
            "conversion": round((total_orders / total_visits) * 100, 2) if total_visits > 0 else 0
        },
        "reputation": {
            "level": reputation.get("level_id", ""),
            "power_seller": rep_data.get("power_seller_status", ""),
            "cancellations": metrics_rep.get("cancellations", {}).get("rate", 0),
            "claims": metrics_rep.get("claims", {}).get("rate", 0),
            "delayed": metrics_rep.get("delayed_handling_time", {}).get("rate", 0),
            "sales_completed": transactions.get("completed", 0),
            "ratings": reputation.get("ratings", {})
        }
    })

@app.route("/api/debug-advertisers/<user_id>")
def debug_advertisers(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    results = {}
    for product_id in ["PADS", "SPA", "PDA", "DISPLAY"]:
        r = requests.get(
            "https://api.mercadolibre.com/advertising/advertisers",
            params={"product_id": product_id},
            headers={"Authorization": "Bearer " + token, "Api-Version": "1"}
        )
        results[product_id] = {"status": r.status_code, "response": r.json() if r.text else {}}

    r_all = requests.get(
        "https://api.mercadolibre.com/advertising/advertisers",
        headers={"Authorization": "Bearer " + token, "Api-Version": "1"}
    )
    results["sem_product_id"] = {"status": r_all.status_code, "response": r_all.json()}

    return jsonify({
        "user_id": user_id,
        "nickname": seller["nickname"],
        "advertiser_lookup": results
    })

@app.route("/api/debug/<user_id>")
def debug_ads(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    urls = [
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns",
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/campaigns",
        "https://api.mercadolibre.com/advertising/advertisers/" + user_id + "/product_ads/campaigns",
        "https://api.mercadolibre.com/advertising/advertisers/" + user_id + "/campaigns",
    ]
    results = {}
    for url in urls:
        r = requests.get(url, headers={"Authorization": "Bearer " + token, "Api-Version": "1"})
        results[url] = {"status": r.status_code, "response": r.json() if r.text else {}}

    return jsonify({
        "user_id": user_id,
        "advertiser_id_found": advertiser_id,
        "endpoints_tested": results
    })

@app.route("/api/debug-metrics/<user_id>")
def debug_metrics(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    r_orders = requests.get(
        "https://api.mercadolibre.com/orders/search",
        params={"seller": user_id, "order.status": "paid", "limit": 1},
        headers={"Authorization": "Bearer " + token}
    )
    r_visits = requests.get(
        "https://api.mercadolibre.com/users/" + user_id + "/items_visits",
        params={"date_from": "2026-05-01", "date_to": "2026-05-26"},
        headers={"Authorization": "Bearer " + token}
    )
    r_rep = requests.get(
        "https://api.mercadolibre.com/users/" + user_id,
        headers={"Authorization": "Bearer " + token}
    )
    r_promos1 = requests.get(
        "https://api.mercadolibre.com/promotions/search",
        params={"seller_id": user_id, "type": "DEAL", "status": "started"},
        headers={"Authorization": "Bearer " + token}
    )
    r_promos2 = requests.get(
        "https://api.mercadolibre.com/seller-promotions/users/" + user_id + "/promotions",
        headers={"Authorization": "Bearer " + token}
    )
    r_promos3 = requests.get(
        "https://api.mercadolibre.com/users/" + user_id + "/deals",
        headers={"Authorization": "Bearer " + token}
    )

    def safe_json(r):
        try:
            return r.json() if r.text and r.text.strip() else {}
        except Exception:
            return {"raw": r.text[:200] if r.text else ""}

    return jsonify({
        "orders":           {"status": r_orders.status_code,  "response": safe_json(r_orders)},
        "visits":           {"status": r_visits.status_code,  "response": safe_json(r_visits)},
        "reputation":       {"status": r_rep.status_code,     "response": safe_json(r_rep)},
        "promotions_search":{"status": r_promos1.status_code, "response": safe_json(r_promos1)},
        "seller_promotions":{"status": r_promos2.status_code, "response": safe_json(r_promos2)},
        "deals":            {"status": r_promos3.status_code, "response": safe_json(r_promos3)}
    })


@app.route("/api/ads/<user_id>/campaign/<camp_id>/products")
@login_required
def get_campaign_products(user_id, camp_id):
    """Retorna produtos e metricas individuais de uma campanha."""
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    # Buscar todos os itens do advertiser e filtrar pela campanha
    items = []
    offset = 0
    while True:
        r = requests.get(
            f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items",
            params={"limit": 50, "offset": offset},
            headers={"Authorization": "Bearer " + token, "Api-Version": "1"},
            timeout=10
        )
        if not r.ok: break
        try:
            d = r.json()
        except: break
        results = d.get("results", [])
        if not results: break
        for item in results:
            if str(item.get("campaign_id", "")) == str(camp_id):
                items.append(item)
        total = d.get("paging", {}).get("total", 0)
        offset += len(results)
        if offset >= total or len(results) == 0: break

    if not items:
        return jsonify({"products": [], "total": 0})

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def fetch_item_metrics(item):
        item_id   = item.get("item_id", "")
        title     = item.get("title", str(item_id))
        status    = item.get("status", "active")
        price     = item.get("price", 0)
        thumbnail = item.get("thumbnail", "")

        m = {}
        if date_from and date_to and item_id:
            # Endpoint confirmado funcionando: /advertising/MLB/product_ads/items/{item_id}
            try:
                rm = requests.get(
                    f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
                    params={
                        "date_from": date_from, "date_to": date_to,
                        "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"
                    },
                    headers={"Authorization": "Bearer " + token, "Api-Version": "2"}, timeout=8
                )
                if rm.ok and rm.text:
                    rd = rm.json()
                    if isinstance(rd, dict) and "metrics" in rd:
                        m = rd["metrics"]
            except: pass

        spend    = m.get("cost", 0)
        revenue  = m.get("total_amount", m.get("direct_amount", 0))
        clicks   = m.get("clicks", 0)
        imps     = m.get("prints", m.get("impressions", 0))
        direct   = m.get("direct_amount", 0)
        indirect = m.get("indirect_amount", 0)
        orders   = m.get("advertising_items_quantity", m.get("direct_items_quantity", 0))
        direct_orders   = m.get("direct_items_quantity", 0)
        indirect_orders = m.get("indirect_items_quantity", 0)
        cvr      = m.get("cvr", 0)

        return {
            "item_id":   str(item_id),
            "title":     title,
            "status":    status,
            "price":     price,
            "thumbnail": thumbnail,
            "spend":     round(spend, 2),
            "revenue":   round(revenue, 2),
            "direct_revenue":   round(direct, 2),
            "indirect_revenue": round(indirect, 2),
            "clicks":    clicks,
            "impressions": imps,
            "orders":    orders,
            "direct_orders":   direct_orders,
            "indirect_orders": indirect_orders,
            "cvr":       round(cvr, 2) if cvr else 0,
            "roas":      round(revenue / spend, 2) if spend > 0 else 0,
            "acos":      round((spend / revenue) * 100, 1) if revenue > 0 else 0,
            "cpc":       round(spend / clicks, 2) if clicks > 0 else 0,
            "ctr":       round((clicks / imps) * 100, 2) if imps > 0 else 0,
        }

    products = []
    with ThreadPoolExecutor(max_workers=6) as ex:
        futures = {ex.submit(fetch_item_metrics, item): item for item in items[:20]}
        for f in as_completed(futures):
            try:
                result = f.result()
                if result: products.append(result)
            except: pass

    # Agrupar produtos com mesmo user_product_id (mesmo produto, anúncios diferentes)
    # Cada item do array 'items' tem user_product_id — precisamos mapear item_id → user_product_id
    upid_map = {str(item.get("item_id","")): item.get("user_product_id","") for item in items}
    family_map = {str(item.get("item_id","")): item.get("family_name", item.get("title","")) for item in items}

    groups = {}
    for p in products:
        upid = upid_map.get(p["item_id"], "") or p["item_id"]
        if upid not in groups:
            groups[upid] = {
                "group_id":    upid,
                "title":       family_map.get(p["item_id"], p["title"]),
                "items":       [],
                "spend":       0, "revenue": 0, "direct_revenue": 0,
                "indirect_revenue": 0, "clicks": 0, "impressions": 0, "orders": 0,
            }
        groups[upid]["items"].append({
            "item_id":  p["item_id"],
            "title":    p["title"],
            "status":   p["status"],
            "price":    p["price"],
            "thumbnail":p["thumbnail"],
            "spend":    p["spend"],
            "revenue":  p["revenue"],
            "clicks":   p["clicks"],
            "impressions": p["impressions"],
        })
        groups[upid]["spend"]            += p["spend"]
        groups[upid]["revenue"]          += p["revenue"]
        groups[upid]["direct_revenue"]   += p["direct_revenue"]
        groups[upid]["indirect_revenue"] += p["indirect_revenue"]
        groups[upid]["clicks"]           += p["clicks"]
        groups[upid]["impressions"]      += p["impressions"]
        groups[upid]["orders"]           += p["orders"]

    # Recalcular métricas consolidadas
    grouped = []
    for upid, g in groups.items():
        sp = g["spend"]; rv = g["revenue"]; cl = g["clicks"]; im = g["impressions"]
        g["roas"] = round(rv / sp, 2) if sp > 0 else 0
        g["acos"] = round((sp / rv) * 100, 1) if rv > 0 else 0
        g["cpc"]  = round(sp / cl, 2) if cl > 0 else 0
        g["ctr"]  = round((cl / im) * 100, 2) if im > 0 else 0
        g["spend"]   = round(sp, 2)
        g["revenue"] = round(rv, 2)
        grouped.append(g)

    grouped.sort(key=lambda x: x["spend"], reverse=True)
    return jsonify({"products": grouped, "total": len(grouped)})


@app.route("/api/ads/<user_id>/campaign/<camp_id>")
@login_required
def get_campaign_detail(user_id, camp_id):
    """Retorna detalhes completos de uma campanha específica."""
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    # Buscar dados da campanha
    camp_data = {}
    for url in [
        f"https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/{camp_id}",
        f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/campaigns/{camp_id}",
        f"https://api.mercadolibre.com/advertising/advertisers/{aid}/campaigns/{camp_id}",
    ]:
        r = requests.get(url, headers={"Authorization": "Bearer " + token, "Api-Version": "1"}, timeout=8)
        if r.ok and r.text:
            try:
                camp_data = r.json()
                break
            except: pass

    # Buscar métricas detalhadas
    metrics_data = {}
    if date_from and date_to:
        r = requests.get(
            f"https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/{camp_id}",
            params={
                "date_from": date_from,
                "date_to": date_to,
                "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"
            },
            headers={"Authorization": "Bearer " + token, "Api-Version": "2"}, timeout=8
        )
        if r.ok and r.text:
            try:
                metrics_data = r.json()
            except: pass

    # Extrair métricas
    m = {}
    if isinstance(metrics_data, dict) and "metrics" in metrics_data:
        m = metrics_data["metrics"]
    elif isinstance(metrics_data, dict):
        m = metrics_data

    spend   = m.get("cost", 0)
    revenue = m.get("total_amount", m.get("direct_amount", 0))
    clicks  = m.get("clicks", 0)
    imps    = m.get("prints", m.get("impressions", 0))
    direct  = m.get("direct_amount", 0)
    indirect = m.get("indirect_amount", 0)

    return jsonify({
        "id":           camp_id,
        "name":         camp_data.get("name", "Campanha " + str(camp_id)),
        "status":       camp_data.get("status", "unknown"),
        "type":         camp_data.get("type", camp_data.get("campaign_type", "-")),
        "budget":       camp_data.get("budget", camp_data.get("daily_budget", 0)),
        "budget_type":  camp_data.get("budget_type", "daily"),
        "start_date":   camp_data.get("start_date", camp_data.get("start", "")),
        "end_date":     camp_data.get("end_date", camp_data.get("end", "")),
        "spend":        round(spend, 2),
        "revenue":      round(revenue, 2),
        "direct_revenue": round(direct, 2),
        "indirect_revenue": round(indirect, 2),
        "clicks":       clicks,
        "impressions":  imps,
        "roas":         round(revenue / spend, 2) if spend > 0 else 0,
        "acos":         round((spend / revenue) * 100, 1) if revenue > 0 else 0,
        "ctr":          round((clicks / imps) * 100, 2) if imps > 0 else 0,
        "cpc":          round(spend / clicks, 2) if clicks > 0 else 0,
    })


@app.route("/api/ai/analyze-campaign", methods=["POST"])
@login_required
def ai_analyze_campaign():
    """Analisa uma campanha usando a API do Claude com a skill de ADS."""
    import json as json_lib2

    data = request.get_json() or {}
    camp = data.get("camp", {})
    products = data.get("products", [])

    # Montar o contexto da campanha
    date_from = camp.get("_date_from", "")
    date_to   = camp.get("_date_to", "")

    products_text = ""
    if products:
        products_text = "\n\nPRODUTOS NA CAMPANHA:\n"
        for p in products:
            products_text += f"""
- {p.get('title','?')} ({p.get('item_id','')})
  Status: {p.get('status','?')}
  Investimento: R${p.get('spend',0):.2f}
  Receita: R${p.get('revenue',0):.2f}
  ROAS: {p.get('roas',0)}x
  ACOS: {p.get('acos',0)}%
  Impressões: {p.get('impressions',0)}
  Cliques: {p.get('clicks',0)}
  CTR: {p.get('ctr',0)}%
  CPC: R${p.get('cpc',0):.2f}
  Vendas ADS: {p.get('orders',0)}
"""

    prompt = f"""Analise a seguinte campanha de Product Ads do Mercado Livre e forneça um diagnóstico completo com recomendações de ação.

CAMPANHA: {camp.get('name','?')}
PERÍODO: {date_from} a {date_to}
STATUS: {camp.get('status','?')}
TIPO: {camp.get('type','-')}
ORÇAMENTO: {camp.get('budget',0)} ({camp.get('budget_type','daily')})

MÉTRICAS CONSOLIDADAS:
- Investimento: R${camp.get('spend',0):.2f}
- Receita ADS: R${camp.get('revenue',0):.2f}
- Receita Direta: R${camp.get('direct_revenue',0):.2f}
- Receita Indireta: R${camp.get('indirect_revenue',0):.2f}
- ROAS: {camp.get('roas',0)}x
- ACOS: {camp.get('acos',0)}%
- Cliques: {camp.get('clicks',0)}
- Impressões: {camp.get('impressions',0)}
- CTR: {camp.get('ctr',0)}%
- CPC: R${camp.get('cpc',0):.2f}
{products_text}

Forneça a análise em HTML usando exatamente este formato:

<div style="font-family:Inter,sans-serif;color:#f0f0f5">

  <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;padding-bottom:12px;border-bottom:1px solid rgba(255,255,255,.1)">
    <span style="font-size:18px">🤖</span>
    <span style="font-weight:700;font-size:14px">Análise IA — {camp.get('name','?')}</span>
  </div>

  <!-- ZONA DE PERFORMANCE -->
  <div style="background:#26262f;border-radius:10px;padding:14px;margin-bottom:12px;border-left:4px solid [COR_ZONA]">
    <div style="font-size:10px;color:#6e6e88;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">Zona de Performance</div>
    <div style="font-size:20px;font-weight:700">[EMOJI] [NOME_ZONA]</div>
    <div style="font-size:12px;color:#9898b0;margin-top:4px">[DESCRIÇÃO_ZONA]</div>
  </div>

  <!-- DIAGNÓSTICO -->
  <div style="margin-bottom:12px">
    <div style="font-size:11px;font-weight:700;color:#6e6e88;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Diagnóstico</div>
    <div style="font-size:13px;color:#d0d0e0;line-height:1.6">[DIAGNÓSTICO_DETALHADO]</div>
  </div>

  <!-- DECISÃO PRINCIPAL -->
  <div style="background:#1a1a22;border-radius:10px;padding:14px;margin-bottom:12px;border:1px solid rgba(255,255,255,.08)">
    <div style="font-size:11px;font-weight:700;color:#6e6e88;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Decisão Recomendada</div>
    <div style="font-size:14px;font-weight:600;color:#f0f0f5;margin-bottom:6px">[DECISÃO_PRINCIPAL]</div>
    <div style="font-size:12px;color:#9898b0">[INSTRUÇÃO_ESPECÍFICA_COM_VALORES_EXATOS]</div>
  </div>

  <!-- PRÓXIMOS PASSOS -->
  <div style="margin-bottom:12px">
    <div style="font-size:11px;font-weight:700;color:#6e6e88;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Próximos Passos</div>
    [LISTA_DE_PASSOS_EM_HTML]
  </div>

  <!-- ALERTAS SE HOUVER -->
  [ALERTAS_OPCIONAIS]

</div>

Use cores: #2ecc9a (verde/escalar), #f5a623 (âmbar/atenção), #ff5c7a (vermelho/crítico), #4a5fe8 (azul/manter).
Seja objetivo e específico — informe valores exatos de ROAS Objetivo recomendado, porcentagem de aumento de orçamento, etc.
Responda SOMENTE com o HTML, sem texto antes ou depois."""

    system_prompt = """Você é um especialista em Product Ads do Mercado Livre com experiência em 16 sellers e +40 relatórios gerados. 

Você domina completamente a skill de análise de campanhas ADS ML, incluindo:
- Zonas ACOS: Escalar (<10%), Ótimo (10-15%), Saudável (15-22%), Atenção (22-30%), Sangrando (>30%)
- ROAS Objetivo deve ser configurado próximo ao ROAS histórico real
- NUNCA alterar orçamento e ROAS Objetivo ao mesmo tempo
- Campanhas com <30 dias estão em aprendizado — não agir se ACOS < 40%
- Escalar = aumentar orçamento (não ROAS Objetivo)
- Ajustar = aumentar ROAS Objetivo em 0,5x incrementos
- Pausar só em casos extremos (ACOS > 60% ou sem estoque)
- Indiretas são contexto, diretas são base de decisão

Seja preciso, objetivo e sempre informe valores numéricos específicos nas recomendações.
Responda SOMENTE com HTML formatado conforme solicitado."""

    try:
        anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 1500,
                "system": system_prompt,
                "messages": [{"role": "user", "content": prompt}]
            },
            headers={
                "Content-Type": "application/json",
                "anthropic-version": "2023-06-01",
                "x-api-key": anthropic_key
            },
            timeout=30
        )
        if not resp.ok:
            err_detail = resp.text[:500]
            print(f"[AI] Erro {resp.status_code}: {err_detail}")
            return jsonify({"html": f'<div style="color:#ff5c7a;padding:12px">Erro da API: {resp.status_code} — {err_detail}</div>'}), 500

        result = resp.json()
        html = result["content"][0]["text"]
        return jsonify({"html": html})

    except Exception as e:
        print(f"[AI] Exceção: {str(e)}")
        return jsonify({"error": str(e), "html": f'<div style="color:#ff5c7a;padding:12px">Erro: {str(e)}</div>'}), 500


@app.route("/api/debug-item-metrics/<user_id>/<camp_id>/<item_id>")
def debug_item_metrics(user_id, camp_id, item_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id
    date_from = request.args.get("date_from", "2026-06-01")
    date_to   = request.args.get("date_to",   "2026-06-15")

    results = {}
    tests = [
        ("items_with_metrics_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items",
         {"campaign_id": camp_id, "item_id": item_id, "date_from": date_from, "date_to": date_to,
          "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "1"),
        ("items_with_metrics_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items",
         {"campaign_id": camp_id, "item_id": item_id, "date_from": date_from, "date_to": date_to,
          "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("mlb_campaign_item_v2", f"https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/{camp_id}",
         {"item_id": item_id, "date_from": date_from, "date_to": date_to,
          "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("advertiser_campaign_item_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/campaigns/{camp_id}",
         {"item_id": item_id, "date_from": date_from, "date_to": date_to,
          "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("item_summary_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to}, "1"),
        ("item_summary_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to,
          "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("mlb_item_v2", f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to,
          "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("items_no_filter_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items",
         {"item_id": item_id, "date_from": date_from, "date_to": date_to}, "1"),
    ]
    for name, url, params, version in tests:
        try:
            r = requests.get(url, params=params,
                headers={"Authorization": "Bearer " + token, "Api-Version": version}, timeout=8)
            results[name] = {"status": r.status_code, "url": url, "params": params,
                             "response": r.json() if r.text else {}}
        except Exception as e:
            results[name] = {"error": str(e)}

    return jsonify({"aid": aid, "camp_id": camp_id, "item_id": item_id, "results": results})


@app.route("/api/debug-camp-products/<user_id>/<camp_id>")
def debug_camp_products(user_id, camp_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    date_from = request.args.get("date_from", "2026-06-01")
    date_to   = request.args.get("date_to",   "2026-06-15")

    results = {}
    tests = [
        # Relatório por item para o advertiser filtrado por campanha
        ("items_report_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id}, "1"),
        ("items_report_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/items", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id}, "2"),
        # Reports
        ("reports_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/reports", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id, "group_by": "ITEM"}, "1"),
        ("reports_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/reports", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id, "group_by": "ITEM"}, "2"),
        # Endpoint de summary por item
        ("summary_items_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/campaigns/{camp_id}/summary", {"date_from": date_from, "date_to": date_to, "group_by": "ITEM"}, "1"),
        ("summary_items_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/campaigns/{camp_id}/summary", {"date_from": date_from, "date_to": date_to, "group_by": "ITEM"}, "2"),
        # MLB genérico
        ("mlb_items_v2", f"https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/{camp_id}", {"date_from": date_from, "date_to": date_to, "group_by": "ITEM"}, "2"),
        ("mlb_ads_v2", f"https://api.mercadolibre.com/advertising/MLB/product_ads/ads", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id}, "2"),
        # Ads do advertiser
        ("advertiser_ads_v1", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/ads", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id}, "1"),
        ("advertiser_ads_v2", f"https://api.mercadolibre.com/advertising/advertisers/{aid}/product_ads/ads", {"date_from": date_from, "date_to": date_to, "campaign_id": camp_id}, "2"),
    ]
    for name, url, params, version in tests:
        try:
            r = requests.get(url, params=params, headers={"Authorization": "Bearer " + token, "Api-Version": version}, timeout=8)
            results[name] = {"status": r.status_code, "url": url, "response": r.json() if r.text else {}}
        except Exception as e:
            results[name] = {"error": str(e), "url": url}

    return jsonify({"advertiser_id": aid, "camp_id": camp_id, "results": results})


@app.route("/api/debug-campaign/<user_id>/<camp_id>")
def debug_campaign(user_id, camp_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    date_from = request.args.get("date_from", "2026-04-01")
    date_to   = request.args.get("date_to",   "2026-05-26")

    results = {}
    tests = [
        ("MLB_v2_metrics", "https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/" + camp_id, {"date_from": date_from, "date_to": date_to, "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("MLB_v1", "https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/" + camp_id, {"date_from": date_from, "date_to": date_to}, "1"),
        ("aid_v2", "https://api.mercadolibre.com/advertising/" + aid + "/product_ads/campaigns/" + camp_id, {"date_from": date_from, "date_to": date_to, "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("aid_campaigns_v2", "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns/" + camp_id, {"date_from": date_from, "date_to": date_to, "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("MLB_no_params", "https://api.mercadolibre.com/advertising/MLB/product_ads/campaigns/" + camp_id, {}, "2"),
    ]
    for name, url, params, version in tests:
        r = requests.get(url, params=params, headers={"Authorization": "Bearer " + token, "Api-Version": version})
        try:
            results[name] = {"status": r.status_code, "url": url, "response": r.json()}
        except Exception:
            results[name] = {"status": r.status_code, "url": url, "response": r.text[:300]}

    return jsonify({
        "user_id": user_id,
        "advertiser_id": aid,
        "camp_id": camp_id,
        "results": results
    })


@app.route("/api/debug-camp-detail/<user_id>")
def debug_camp_detail(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id

    # Busca campanhas e retorna JSON completo
    r = requests.get(
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns",
        headers={"Authorization": "Bearer " + token, "Api-Version": "1"}
    )
    try:
        camp_data = r.json()
    except Exception:
        camp_data = r.text[:500]

    # Testa endpoint de summary de metricas do advertiser
    r2 = requests.get(
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns/summary",
        params={"date_from": "2026-04-01", "date_to": "2026-05-26"},
        headers={"Authorization": "Bearer " + token, "Api-Version": "1"}
    )
    try:
        summary_data = r2.json()
    except Exception:
        summary_data = r2.text[:500]

    # Testa endpoint de reports
    r3 = requests.get(
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/reports",
        params={"date_from": "2026-04-01", "date_to": "2026-05-26"},
        headers={"Authorization": "Bearer " + token, "Api-Version": "1"}
    )
    try:
        reports_data = r3.json()
    except Exception:
        reports_data = r3.text[:500]

    return jsonify({
        "advertiser_id": aid,
        "campaigns_raw": camp_data,
        "summary_endpoint": {"status": r2.status_code, "response": summary_data},
        "reports_endpoint": {"status": r3.status_code, "response": reports_data}
    })


@app.route("/api/debug-metrics2/<user_id>")
def debug_metrics2(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    advertiser_id = get_advertiser_id(user_id, token)
    aid = advertiser_id if advertiser_id else user_id
    camp_id = "357533684"
    sf_id = "14"
    date_from = "2026-04-01"
    date_to = "2026-05-26"

    urls = [
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns/" + camp_id + "/metrics/days",
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns/" + camp_id + "/daily_metrics",
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/metrics",
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/metrics",
        "https://api.mercadolibre.com/advertising/advertisers/" + aid + "/product_ads/campaigns/" + camp_id + "/clicks",
        "https://api.mercadolibre.com/advertising/" + aid + "/campaigns/" + camp_id + "/metrics",
        "https://api.mercadolibre.com/pads/advertisers/" + aid + "/campaigns/" + camp_id + "/metrics",
    ]

    results = {}
    for url in urls:
        r = requests.get(url, params={"date_from": date_from, "date_to": date_to}, headers={"Authorization": "Bearer " + token, "Api-Version": "1"})
        try:
            results[url] = {"status": r.status_code, "response": r.json()}
        except Exception:
            results[url] = {"status": r.status_code, "response": r.text[:300]}

    return jsonify({"advertiser_id": aid, "camp_id": camp_id, "results": results})


@app.route("/api/debug-sales/<user_id>")
def debug_sales(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    from datetime import date as ddate, timedelta as tdelta
    today = ddate.today()
    date_from = (today - tdelta(days=7)).isoformat()
    date_to   = today.isoformat()

    # Testa diferentes status
    results = {}
    for status in ["paid", "cancelled", "delivered", "confirmed", "payment_required", "payment_in_process"]:
        r = requests.get(
            "https://api.mercadolibre.com/orders/search",
            params={
                "seller": user_id,
                "order.date_created.from": date_from + "T00:00:00.000-03:00",
                "order.date_created.to":   date_to + "T23:59:59.000-03:00",
                "order.status": status,
                "limit": 1,
                "offset": 0
            },
            headers={"Authorization": "Bearer " + token}
        )
        if r.ok and r.text:
            d = r.json()
            results[status] = {
                "total": d.get("paging", {}).get("total", 0),
                "status": r.status_code
            }
        else:
            results[status] = {"total": 0, "status": r.status_code}

    # Busca paid com paginacao e mostra totais
    all_paid = []
    offset = 0
    limit = 50
    total_api = 0
    pages = []
    while offset <= 500:
        r = requests.get(
            "https://api.mercadolibre.com/orders/search",
            params={
                "seller": user_id,
                "order.date_created.from": date_from + "T00:00:00.000-03:00",
                "order.date_created.to":   date_to + "T23:59:59.000-03:00",
                "order.status": "paid",
                "limit": limit,
                "offset": offset,
                "sort": "date_asc"
            },
            headers={"Authorization": "Bearer " + token}
        )
        if not r.ok or not r.text:
            break
        data = r.json()
        results_page = data.get("results", [])
        total_api = data.get("paging", {}).get("total", 0)
        pages.append({"offset": offset, "count": len(results_page), "total_reported": total_api})
        all_paid.extend(results_page)
        offset += limit
        if offset >= total_api or not results_page:
            break

    gmv = sum(order_value(o) for o in all_paid)

    return jsonify({
        "date_range": {"from": date_from, "to": date_to},
        "status_totals": results,
        "paid_pagination": pages,
        "paid_fetched": len(all_paid),
        "paid_total_api": total_api,
        "gmv_calculated": round(gmv, 2),
        "sample_order": all_paid[0] if all_paid else None
    })


@app.route("/api/debug-sales2/<user_id>")
def debug_sales2(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    from datetime import date as ddate, timedelta as tdelta
    today = ddate.today()
    date_from = (today - tdelta(days=30)).isoformat()
    date_to   = today.isoformat()

    # Busca total por cada status individualmente
    statuses = ["paid", "cancelled", "confirmed", "payment_in_process",
                "payment_required", "partially_refunded", "pending_cancel",
                "delivered", "invalid"]
    status_totals = {}
    for st in statuses:
        r = requests.get(
            "https://api.mercadolibre.com/orders/search",
            params={
                "seller": user_id,
                "order.date_created.from": date_from + "T00:00:00.000-03:00",
                "order.date_created.to":   date_to + "T23:59:59.000-03:00",
                "order.status": st,
                "limit": 1, "offset": 0
            },
            headers={"Authorization": "Bearer " + token}
        )
        if r.ok and r.text:
            try:
                status_totals[st] = r.json().get("paging", {}).get("total", 0)
            except Exception:
                status_totals[st] = 0
        else:
            status_totals[st] = {"error": r.status_code}

    # Total sem filtro
    r_all = requests.get(
        "https://api.mercadolibre.com/orders/search",
        params={
            "seller": user_id,
            "order.date_created.from": date_from + "T00:00:00.000-03:00",
            "order.date_created.to":   date_to + "T23:59:59.000-03:00",
            "limit": 1, "offset": 0
        },
        headers={"Authorization": "Bearer " + token}
    )
    total_no_filter = r_all.json().get("paging", {}).get("total", 0) if r_all.ok and r_all.text else 0

    return jsonify({
        "date_range": {"from": date_from, "to": date_to},
        "total_no_filter": total_no_filter,
        "by_status": status_totals,
        "sum_known_statuses": sum(v for v in status_totals.values() if isinstance(v, int))
    })


@app.route("/api/debug-reports/<user_id>")
def debug_reports(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    from datetime import date as ddate, timedelta as tdelta
    today = ddate.today()
    date_from = (today - tdelta(days=30)).isoformat()
    date_to   = today.isoformat()

    results = {}

    # Endpoint 1: data-reporting/seller_performance
    r1 = requests.get(
        "https://api.mercadolibre.com/data-reporting/seller_performance",
        params={"caller.id": user_id, "date_from": date_from, "date_to": date_to},
        headers={"Authorization": "Bearer " + token}
    )
    results["seller_performance"] = {"status": r1.status_code, "response": r1.json() if r1.ok and r1.text else r1.text[:200]}

    # Endpoint 2: seller_performance/search
    r2 = requests.get(
        "https://api.mercadolibre.com/data-reporting/seller_performance/search",
        params={"caller.id": user_id, "date_from": date_from, "date_to": date_to},
        headers={"Authorization": "Bearer " + token}
    )
    results["seller_performance_search"] = {"status": r2.status_code, "response": r2.json() if r2.ok and r2.text else r2.text[:200]}

    # Endpoint 3: users/{id}/orders_summary
    r3 = requests.get(
        "https://api.mercadolibre.com/users/" + user_id + "/orders_summary",
        params={"date_from": date_from, "date_to": date_to},
        headers={"Authorization": "Bearer " + token}
    )
    results["orders_summary"] = {"status": r3.status_code, "response": r3.json() if r3.ok and r3.text else r3.text[:200]}

    # Endpoint 4: billing/integration/invoices
    r4 = requests.get(
        "https://api.mercadolibre.com/users/" + user_id + "/sales",
        params={"date_from": date_from, "date_to": date_to},
        headers={"Authorization": "Bearer " + token}
    )
    results["user_sales"] = {"status": r4.status_code, "response": r4.json() if r4.ok and r4.text else r4.text[:200]}

    # Endpoint 5: seller_performance via meli API
    r5 = requests.get(
        "https://api.mercadolibre.com/users/" + user_id + "/items_visits/time_window",
        params={"last": "30", "unit": "day", "ending": date_to},
        headers={"Authorization": "Bearer " + token}
    )
    results["items_visits_window"] = {"status": r5.status_code, "response": r5.json() if r5.ok and r5.text else r5.text[:200]}

    # Endpoint 6: highlights
    r6 = requests.get(
        "https://api.mercadolibre.com/highlights/MLB/seller/" + user_id,
        headers={"Authorization": "Bearer " + token}
    )
    results["highlights"] = {"status": r6.status_code, "response": r6.json() if r6.ok and r6.text else r6.text[:200]}

    # Endpoint 7: v2 orders com status=all
    r7 = requests.get(
        "https://api.mercadolibre.com/orders/search",
        params={
            "seller": user_id,
            "order.date_created.from": date_from + "T00:00:00.000-03:00",
            "order.date_created.to":   date_to + "T23:59:59.000-03:00",
            "limit": 1,
            "offset": 0
        },
        headers={"Authorization": "Bearer " + token}
    )
    results["orders_no_status_filter"] = {
        "status": r7.status_code,
        "total": r7.json().get("paging", {}).get("total", 0) if r7.ok and r7.text else 0
    }

    return jsonify({
        "user_id": user_id,
        "date_range": {"from": date_from, "to": date_to},
        "results": results
    })


@app.route("/api/debug-promos/<user_id>")
def debug_promos(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    results = {}

    endpoints = [
        ("promotions_v1",        "https://api.mercadolibre.com/promotions",
         {"seller_id": user_id}),
        ("seller_deals",         "https://api.mercadolibre.com/seller-promotions/users/" + user_id + "/promotions",
         {}),
        ("seller_deals_v2",      "https://api.mercadolibre.com/seller-promotions/users/" + user_id + "/promotions?limit=10",
         {}),
        ("discount_campaigns",   "https://api.mercadolibre.com/discount-campaigns/users/" + user_id,
         {}),
        ("coupons",              "https://api.mercadolibre.com/users/" + user_id + "/coupons",
         {}),
        ("item_promotions",      "https://api.mercadolibre.com/users/" + user_id + "/item_promotions",
         {}),
        ("promotions_search2",   "https://api.mercadolibre.com/promotions/search",
         {"seller_id": user_id}),
        ("loyalty",              "https://api.mercadolibre.com/loyalty/users/" + user_id + "/summary",
         {}),
        ("mshops_promotions",    "https://api.mercadolibre.com/mshops/promotion-campaigns/seller/" + user_id,
         {}),
        ("price_campaigns",      "https://api.mercadolibre.com/price-campaigns/seller/" + user_id,
         {}),
    ]

    for name, url, params in endpoints:
        try:
            r = requests.get(url, params=params, headers={"Authorization": "Bearer " + token})
            try:
                body = r.json()
            except Exception:
                body = r.text[:300]
            results[name] = {"status": r.status_code, "response": body}
        except Exception as e:
            results[name] = {"error": str(e)}

    return jsonify({"user_id": user_id, "results": results})

@app.route("/notifications", methods=["GET", "POST"])
def notifications():
    # GET: validação do webhook pelo ML (envia ?challenge=xxx, devemos retornar o mesmo valor)
    challenge = request.args.get("challenge")
    if challenge:
        return jsonify({"challenge": challenge}), 200

    # POST: notificação real do ML
    try:
        payload = request.get_json(silent=True) or {}
        topic   = payload.get("topic", "")
        res_id  = payload.get("resource", "")
        user_id = str(payload.get("user_id", ""))
        print(f"[NOTIFICATION] topic={topic} resource={res_id} user_id={user_id}")
    except Exception as e:
        print(f"[NOTIFICATION][ERRO] {e}")

    return "", 200


@app.route("/api/questions/<user_id>")
def get_questions(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    r_all = requests.get(
        "https://api.mercadolibre.com/questions/search",
        params={"seller_id": user_id, "limit": 50, "sort_fields": "date_created", "sort_types": "DESC"},
        headers={"Authorization": "Bearer " + token}
    )
    all_data = r_all.json() if r_all.ok and r_all.text else {}

    r_unans = requests.get(
        "https://api.mercadolibre.com/questions/search",
        params={"seller_id": user_id, "status": "UNANSWERED", "limit": 50},
        headers={"Authorization": "Bearer " + token}
    )
    unans_data = r_unans.json() if r_unans.ok and r_unans.text else {}

    questions = all_data.get("questions", [])
    result = []
    for q in questions:
        result.append({
            "id":           q.get("id"),
            "text":         q.get("text", ""),
            "status":       q.get("status", ""),
            "date_created": q.get("date_created", "")[:16].replace("T", " "),
            "item_id":      q.get("item_id", ""),
            "item_title":   q.get("item_title", ""),
            "answer":       q.get("answer", {}).get("text", "") if q.get("answer") else "",
            "answer_date":  q.get("answer", {}).get("date_created", "")[:16].replace("T", " ") if q.get("answer") else "",
        })

    return jsonify({
        "seller_id":        user_id,
        "nickname":         seller["nickname"],
        "total":            all_data.get("total", 0),
        "total_unanswered": unans_data.get("total", 0),
        "questions":        result
    })


@app.route("/api/debug-financial/<user_id>")
def debug_financial(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    def safe_json(r):
        try: return r.json() if r.text and r.text.strip() else {}
        except: return {"raw": r.text[:300] if r.text else ""}

    results = {}
    endpoints = {
        "balance":           f"https://api.mercadolibre.com/users/{user_id}/mercadopago_account/balance",
        "money_in_accounts": f"https://api.mercadolibre.com/users/{user_id}/money_in_accounts",
        "movements":         f"https://api.mercadolibre.com/users/{user_id}/movements",
        "available_balance": f"https://api.mercadolibre.com/users/{user_id}/available_balance",
        "seller_wallet":     f"https://api.mercadolibre.com/seller-account/{user_id}/balance",
    }
    for name, url in endpoints.items():
        r = requests.get(url, headers={"Authorization": "Bearer " + token}, timeout=8)
        results[name] = {"status": r.status_code, "response": str(safe_json(r))[:200]}
    return jsonify({"user_id": user_id, "results": results})


@app.route("/api/debug-messages/<user_id>")
def debug_messages(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    def safe_json(r):
        try: return r.json() if r.text and r.text.strip() else {}
        except: return {"raw": r.text[:300] if r.text else ""}

    results = {}
    endpoints = {
        "questions":            f"https://api.mercadolibre.com/questions/search?seller_id={user_id}&limit=5",
        "questions_unanswered": f"https://api.mercadolibre.com/questions/search?seller_id={user_id}&status=UNANSWERED&limit=5",
        "messages_unread":      f"https://api.mercadolibre.com/messages/unread?user_id={user_id}",
        "claims":               f"https://api.mercadolibre.com/post-purchase/v1/claims/search?seller_id={user_id}&limit=5",
    }
    for name, url in endpoints.items():
        r = requests.get(url, headers={"Authorization": "Bearer " + token}, timeout=8)
        results[name] = {"status": r.status_code, "response": str(safe_json(r))[:300]}
    return jsonify({"user_id": user_id, "results": results})


@app.route("/api/debug-shipments/<user_id>")
def debug_shipments(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    def safe_json(r):
        try: return r.json() if r.text and r.text.strip() else {}
        except: return {"raw": r.text[:300] if r.text else ""}

    results = {}
    endpoints = {
        "shipments_search":      f"https://api.mercadolibre.com/shipments/search?seller_id={user_id}&limit=5",
        "shipments_by_seller":   f"https://api.mercadolibre.com/users/{user_id}/shipments?limit=5",
        "orders_with_shipments": f"https://api.mercadolibre.com/orders/search?seller={user_id}&limit=3",
        "flex_shipments":        f"https://api.mercadolibre.com/users/{user_id}/flex_handshakes?limit=5",
    }
    for name, url in endpoints.items():
        r = requests.get(url, headers={"Authorization": "Bearer " + token}, timeout=8)
        results[name] = {"status": r.status_code, "response": str(safe_json(r))[:300]}
    return jsonify({"user_id": user_id, "results": results})


@app.route("/api/validate/<user_id>")
def validate(user_id):
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao encontrado"}), 404

    results = {}
    today = datetime.now(timezone.utc).date().isoformat()
    date_from = (datetime.now(timezone.utc).date() - timedelta(days=7)).isoformat()

    def check(name, url, params=None):
        try:
            r = requests.get(url, params=params, headers={"Authorization": "Bearer " + token}, timeout=10)
            data = r.json() if r.text and r.text.strip() else {}
            results[name] = {"status": r.status_code, "ok": r.status_code==200, "preview": str(data)[:120]}
        except Exception as e:
            results[name] = {"status": 0, "ok": False, "preview": str(e)[:120]}

    check("perfil",    f"https://api.mercadolibre.com/users/{user_id}")
    check("vendas",    "https://api.mercadolibre.com/orders/search", {"seller": user_id, "limit": 1})
    check("visitas",   f"https://api.mercadolibre.com/users/{user_id}/items_visits", {"date_from": date_from, "date_to": today})
    check("reputacao", f"https://api.mercadolibre.com/users/{user_id}/seller_reputation")
    check("promocoes", f"https://api.mercadolibre.com/seller-promotions/users/{user_id}", {"app_version": "v2"})
    check("perguntas", f"https://api.mercadolibre.com/questions/search", {"seller_id": user_id, "limit": 1})

    r_adv = requests.get("https://api.mercadolibre.com/advertising/advertisers",
                         params={"product_id": "PADS"},
                         headers={"Authorization": "Bearer " + token, "Api-Version": "1"}, timeout=10)
    adv_data = r_adv.json() if r_adv.ok and r_adv.text else []
    adv_id = adv_data[0]["id"] if isinstance(adv_data, list) and adv_data else None
    results["ads_advertiser"] = {"status": r_adv.status_code, "ok": r_adv.status_code==200, "advertiser_id": adv_id}

    if adv_id:
        check("ads_campanhas", f"https://api.mercadolibre.com/advertising/advertisers/{adv_id}/product_ads/campaigns")
    else:
        results["ads_campanhas"] = {"status": 0, "ok": False, "preview": "advertiser_id nao encontrado"}

    ok_count = sum(1 for v in results.values() if v.get("ok"))
    return jsonify({"seller": seller["nickname"], "user_id": user_id, "score": f"{ok_count}/{len(results)}", "results": results})


@app.route("/api/promo-items/<user_id>/<promo_id>")
def get_promo_items(user_id, promo_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    promo_type = request.args.get("type", "SMART")
    start_date = request.args.get("start_date", "")

    # Busca itens da promoção
    r = requests.get(
        f"https://api.mercadolibre.com/seller-promotions/promotions/{promo_id}/items",
        params={"promotion_type": promo_type, "app_version": "v2"},
        headers={"Authorization": "Bearer " + token}
    )
    if not r.ok:
        return jsonify({"error": "Nao foi possivel buscar itens", "status": r.status_code}), 400

    data = r.json()
    items = data.get("results", [])

    if not items:
        return jsonify({"promo_id": promo_id, "items": [], "total": 0})

    # Para cada item, busca vendas desde o início da promoção
    date_from = start_date[:10] if start_date else datetime.now(timezone.utc).date().isoformat()
    date_to   = datetime.now(timezone.utc).date().isoformat()

    # Busca pedidos no período da promoção
    all_orders = []
    offset = 0
    while offset < 2000:
        r_ord = requests.get(
            "https://api.mercadolibre.com/orders/search",
            params={
                "seller": user_id,
                "order.date_created.from": date_from + "T00:00:00.000-03:00",
                "order.date_created.to":   date_to   + "T23:59:59.000-03:00",
                "limit": 50, "offset": offset, "sort": "date_asc"
            },
            headers={"Authorization": "Bearer " + token}
        )
        if not r_ord.ok: break
        batch = r_ord.json().get("results", [])
        if not batch: break
        all_orders.extend(batch)
        if len(batch) < 50: break
        offset += 50

    SALE_STATUSES = {"confirmed","payment_required","payment_in_process","paid","partially_refunded","partially_paid","pending_cancel","delivered"}
    sale_orders = [o for o in all_orders if o.get("status") in SALE_STATUSES]

    # Agrupa vendas por item_id
    sales_by_item = {}
    for o in sale_orders:
        for oi in o.get("order_items", []):
            iid = str(oi.get("item", {}).get("id", ""))
            qty = oi.get("quantity", 1)
            price = oi.get("unit_price", 0)
            if iid not in sales_by_item:
                sales_by_item[iid] = {"qty": 0, "revenue": 0.0}
            sales_by_item[iid]["qty"]     += qty
            sales_by_item[iid]["revenue"] += qty * price

    # Busca títulos dos itens em batch
    item_ids = [str(item.get("id","")) for item in items if item.get("id")]
    titles = {}
    if item_ids:
        chunk = ",".join(item_ids[:20])
        r_titles = requests.get(
            "https://api.mercadolibre.com/items",
            params={"ids": chunk, "attributes": "id,title"},
            headers={"Authorization": "Bearer " + token}
        )
        if r_titles.ok and r_titles.text:
            for entry in r_titles.json():
                body = entry.get("body", {})
                titles[str(body.get("id",""))] = body.get("title","")

    # Monta resultado
    result = []
    for item in items:
        iid    = str(item.get("id", ""))
        s      = sales_by_item.get(iid, {"qty": 0, "revenue": 0.0})
        status = item.get("status", "")
        orig   = item.get("original_price", 0) or 0

        # seller_percentage:
        # - started: vem direto do campo ou calculado pelo preço atual
        # - candidate: ML não retorna o campo - calcula via suggested_discounted_price
        seller_pct = item.get("seller_percentage", 0)
        if not seller_pct and orig > 0:
            if status == "candidate":
                suggested = item.get("suggested_discounted_price", 0) or 0
                if suggested > 0:
                    seller_pct = round((orig - suggested) / orig * 100, 1)
            elif status == "started":
                price = item.get("price", 0) or 0
                if price > 0:
                    seller_pct = round((orig - price) / orig * 100, 1)

        result.append({
            "item_id":          iid,
            "title":            titles.get(iid, iid),
            "status":           status,
            "price":            item.get("price", 0),
            "original_price":   orig,
            "meli_percentage":  item.get("meli_percentage", 0),
            "seller_percentage": seller_pct,
            "suggested_discounted_price": item.get("suggested_discounted_price", 0),
            "min_discounted_price": item.get("min_discounted_price", 0),
            "start_date":       str(item.get("start_date",""))[:10],
            "end_date":         str(item.get("end_date",""))[:10],
            "qty_sold":         s["qty"],
            "revenue":          round(s["revenue"], 2),
        })

    return jsonify({
        "promo_id":   promo_id,
        "date_from":  date_from,
        "date_to":    date_to,
        "total":      len(result),
        "items":      result
    })



@app.route("/api/item-visits/<user_id>/<item_id>")
def get_item_visits(user_id, item_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")

    if not date_from or not date_to:
        today    = datetime.now(timezone.utc).date()
        date_to  = today.isoformat()
        date_from = (today - tdelta(days=29)).isoformat()

    try:
        r = requests.get(
            f"https://api.mercadolibre.com/users/{user_id}/items_visits",
            params={"ids": item_id, "date_from": date_from, "date_to": date_to},
            headers={"Authorization": "Bearer " + token},
            timeout=8
        )
        if not r.ok:
            return jsonify({"visits": 0, "error": r.status_code})

        data = r.json()
        # Resposta pode ser lista ou dict com data_points
        visits = 0
        if isinstance(data, list):
            for entry in data:
                if str(entry.get("item_id", "")) == str(item_id):
                    visits = entry.get("total_visits", entry.get("visits", 0))
        elif isinstance(data, dict):
            visits = data.get("total_visits", data.get("visits", 0))
            # Pode vir como lista de items dentro do dict
            items_list = data.get("items", data.get("results", []))
            if items_list:
                for entry in items_list:
                    if str(entry.get("item_id", "")) == str(item_id):
                        visits = entry.get("total_visits", entry.get("visits", 0))

        return jsonify({
            "item_id":   item_id,
            "visits":    visits,
            "date_from": date_from,
            "date_to":   date_to
        })
    except Exception as e:
        return jsonify({"visits": 0, "error": str(e)})


@app.route("/api/price-suggestion/<user_id>/<item_id>")
def get_price_suggestion(user_id, item_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    try:
        r = requests.get(
            f"https://api.mercadolibre.com/items/{item_id}/price_suggestion",
            headers={"Authorization": "Bearer " + token},
            timeout=8
        )
        if not r.ok:
            return jsonify({"error": r.status_code, "available": False})

        data = r.json()

        # Extrai campos principais
        suggested     = data.get("suggested_price", data.get("price_suggestion", {}).get("price", 0)) or 0
        price_min     = data.get("min_price", data.get("price_range", {}).get("min", 0)) or 0
        price_max     = data.get("max_price", data.get("price_range", {}).get("max", 0)) or 0
        median        = data.get("median_price", data.get("price_suggestion", {}).get("median_price", 0)) or 0

        return jsonify({
            "item_id":       item_id,
            "available":     True,
            "suggested":     round(suggested, 2),
            "price_min":     round(price_min, 2),
            "price_max":     round(price_max, 2),
            "median":        round(median, 2),
            "raw":           data  # para debug
        })

    except Exception as e:
        return jsonify({"error": str(e), "available": False})


@app.route("/api/full-stock/<user_id>")
@login_required
def get_full_stock(user_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    """Retorna apenas produtos Full ML com estoque atual - endpoint leve para alertas"""
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404
    try:
        # Busca todos os item_ids do seller
        all_ids = []
        scroll_id = None
        for _ in range(10):
            params = {"limit": 100}
            if scroll_id:
                params["scroll_id"] = scroll_id
            r = requests.get(
                f"https://api.mercadolibre.com/users/{user_id}/items/search",
                params=params,
                headers={"Authorization": "Bearer " + token},
                timeout=8
            )
            if not r.ok: break
            d     = r.json()
            batch = d.get("results", [])
            if not batch: break
            all_ids.extend(batch)
            scroll_id = d.get("scroll_id")
            if not scroll_id or len(batch) < 100: break

        # Busca atributos relevantes em chunks
        full_products = []
        for i in range(0, len(all_ids), 20):
            chunk = all_ids[i:i+20]
            r2 = requests.get(
                "https://api.mercadolibre.com/items",
                params={"ids": ",".join(chunk), "attributes": "id,title,available_quantity,fulfillment,shipping,status"},
                headers={"Authorization": "Bearer " + token},
                timeout=8
            )
            if not r2.ok: continue
            for entry in r2.json():
                body = entry.get("body", {})
                if body.get("status") not in ("active", "paused"): continue
                fulfillment = body.get("fulfillment") or {}
                shipping    = body.get("shipping") or {}
                is_full = bool(
                    fulfillment.get("fulfillment_id") or
                    fulfillment.get("status") == "active" or
                    shipping.get("fulfillment") or
                    shipping.get("logistic_type") == "fulfillment" or
                    shipping.get("logistic_type") == "meli_fulfillment"
                )
                if not is_full: continue
                full_products.append({
                    "item_id": str(body.get("id", "")),
                    "title":   body.get("title", ""),
                    "stock":   body.get("available_quantity", 0) or 0,
                    "status":  body.get("status", "active"),
                })

        return jsonify({"user_id": user_id, "products": full_products})
    except Exception as e:
        return jsonify({"error": str(e), "products": []}), 500


@app.route("/api/item-promos/<user_id>/<item_id>")
def get_item_promos(user_id, item_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    try:
        # Busca promoções usando a mesma lógica que funciona em get_promotions
        all_promos = []
        r1 = requests.get(
            f"https://api.mercadolibre.com/seller-promotions/users/{user_id}",
            params={"app_version": "v2"},
            headers={"Authorization": "Bearer " + token},
            timeout=8
        )
        if r1.ok and r1.text:
            d = r1.json()
            raw = d if isinstance(d, list) else d.get("results", d.get("promotions", []))
            all_promos = [p for p in raw if p.get("status") in ("started", "active")]

        if not all_promos:
            r2 = requests.get(
                f"https://api.mercadolibre.com/seller-promotions/users/{user_id}/promotions",
                params={"app_version": "v2"},
                headers={"Authorization": "Bearer " + token},
                timeout=8
            )
            if r2.ok and r2.text:
                d = r2.json()
                raw = d if isinstance(d, list) else d.get("results", d.get("promotions", []))
                all_promos = [p for p in raw if p.get("status") in ("started", "active")]

        if not all_promos:
            return jsonify({"promos": []})
        type_map   = {"SMART": "Smart", "DEAL": "Deal", "PRICE_MATCHING_MELI_ALL": "Price Match",
                      "LIGHTNING": "Relâmpago", "SELLER_CAMPAIGN": "Campanha"}
        result = []

        for promo in all_promos:
            promo_id   = promo.get("id", "")
            promo_type = promo.get("type", "SMART")
            if not promo_id:
                continue

            # Busca itens dessa promoção para ver se o item está incluído
            r_items = requests.get(
                f"https://api.mercadolibre.com/seller-promotions/promotions/{promo_id}/items",
                params={"promotion_type": promo_type, "app_version": "v2"},
                headers={"Authorization": "Bearer " + token},
                timeout=8
            )
            if not r_items.ok:
                continue

            items_data = r_items.json()
            items_list = items_data.get("results", items_data.get("items", [])) if isinstance(items_data, dict) else []

            for it in items_list:
                if str(it.get("id", "")) != str(item_id):
                    continue

                status   = it.get("status", "")
                orig     = it.get("original_price", 0) or 0
                price    = it.get("price", 0) or 0
                sugg     = it.get("suggested_discounted_price", 0) or 0

                # Calcula desconto
                if status == "started" and orig > 0 and price > 0:
                    discount_pct = round((orig - price) / orig * 100, 1)
                elif status == "candidate" and orig > 0 and sugg > 0:
                    discount_pct = round((orig - sugg) / orig * 100, 1)
                else:
                    discount_pct = it.get("seller_percentage", 0) or 0

                result.append({
                    "promo_id":    promo_id,
                    "name":        promo.get("name", promo_id),
                    "type":        type_map.get(promo_type, promo_type),
                    "status":      status,
                    "discount_pct": discount_pct,
                    "start_date":  str(promo.get("start_date", ""))[:10],
                    "end_date":    str(promo.get("end_date", promo.get("finish_date", promo.get("to_date", promo.get("to", "")))))[:10],
                    "original_price": orig,
                    "promo_price":    price if status == "started" else sugg,
                })
                break  # achou o item nessa promoção, próxima promoção

        return jsonify({"item_id": item_id, "promos": result})

    except Exception as e:
        return jsonify({"promos": [], "error": str(e)})


@app.route("/api/debug-promo-items/<user_id>/<promo_id>")
def debug_promo_items(user_id, promo_id):
    try:
        token, seller = get_seller_token(user_id)
        if not seller:
            return jsonify({"error": "Seller nao encontrado"}), 404

        results = {}
        for ptype in ["SMART", "DEAL", "PRICE_MATCHING_MELI_ALL", "LIGHTNING", "SELLER_CAMPAIGN"]:
            try:
                url = f"https://api.mercadolibre.com/seller-promotions/promotions/{promo_id}/items"
                r = requests.get(url,
                    params={"promotion_type": ptype, "app_version": "v2"},
                    headers={"Authorization": "Bearer " + token}, timeout=8)
                try:
                    raw = r.json()
                except Exception:
                    raw = {"raw_text": r.text[:500]}

                items_sample = []
                if isinstance(raw, dict):
                    items_sample = raw.get("results", raw.get("items", []))[:2]
                elif isinstance(raw, list):
                    items_sample = raw[:2]

                results[ptype] = {
                    "http_status": r.status_code,
                    "response_keys": list(raw.keys()) if isinstance(raw, dict) else [],
                    "first_items_raw": items_sample,
                    "raw_preview": str(raw)[:800]
                }
            except Exception as e:
                results[ptype] = {"error": str(e)}

        return jsonify({"promo_id": promo_id, "user_id": user_id, "results": results})
    except Exception as e:
        return jsonify({"fatal_error": str(e)}), 500


@app.route("/api/debug-item/<user_id>/<item_id>")
def debug_item(user_id, item_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    token, seller = get_seller_token(user_id)
    if not seller:
        return jsonify({"error": "Seller nao autorizado"}), 404

    date_from = request.args.get("date_from", "2026-06-01")
    date_to   = request.args.get("date_to",   "2026-06-15")

    results = {}
    # Testa várias combinações de métricas e versões para descobrir o que a API aceita
    tests = [
        ("v2_padrao",     f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to, "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "2"),
        ("v2_com_orders", f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to, "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,orders"}, "2"),
        ("v2_sem_metrics",f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to}, "2"),
        ("v1_padrao",     f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to, "metrics": "clicks,prints,cost,direct_amount,indirect_amount,total_amount,advertising_items_quantity,direct_items_quantity,indirect_items_quantity,cvr"}, "1"),
        ("v1_sem_metrics",f"https://api.mercadolibre.com/advertising/MLB/product_ads/items/{item_id}",
         {"date_from": date_from, "date_to": date_to}, "1"),
    ]
    for name, url, params, version in tests:
        try:
            r = requests.get(url, params=params,
                             headers={"Authorization": "Bearer " + token, "Api-Version": version},
                             timeout=8)
            try:
                results[name] = {"status": r.status_code, "response": r.json(), "url": url, "params": params, "version": version}
            except:
                results[name] = {"status": r.status_code, "response": r.text[:300], "url": url}
        except Exception as e:
            results[name] = {"error": str(e)}

    return jsonify({"user_id": user_id, "item_id": item_id, "results": results})



# ─── HISTÓRICO DE CAMPANHAS ───────────────────────────────────────────

def collect_campaign_history():
    """Coleta métricas do dia anterior para todas as campanhas de todos os sellers."""
    print("[HISTORY] Iniciando coleta diária de histórico de campanhas...")
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT user_id, access_token, refresh_token, nickname FROM sellers")
        sellers = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        print("[HISTORY][ERRO] Falha ao buscar sellers:", e)
        return

    for seller in sellers:
        user_id = seller["user_id"]
        token   = seller["access_token"]
        try:
            token = refresh_token_if_needed(user_id, token, seller.get("refresh_token",""), seller.get("updated_at", datetime.now(timezone.utc)))
        except: pass

        try:
            campaigns, base_url, aid = get_campaigns(user_id, token)
        except Exception as e:
            print(f"[HISTORY][ERRO] seller {user_id} campanhas:", e)
            continue

        for camp in campaigns:
            camp_id   = str(camp.get("id",""))
            camp_name = camp.get("name","")
            try:
                metrics = get_campaign_metrics(aid, camp_id, token, yesterday, yesterday, base_url or "")
                if isinstance(metrics, dict) and "metrics" in metrics:
                    m = metrics["metrics"]
                elif isinstance(metrics, dict):
                    m = metrics
                else:
                    continue

                spend   = float(m.get("cost", 0))
                revenue = float(m.get("total_amount", m.get("direct_amount", 0)))
                clicks  = int(m.get("clicks", 0))
                imps    = int(m.get("prints", m.get("impressions", 0)))
                orders  = int(m.get("advertising_items_quantity", m.get("direct_items_quantity", 0)))
                cvr     = float(m.get("cvr", 0))
                roas    = round(revenue / spend, 2) if spend > 0 else 0
                acos    = round((spend / revenue) * 100, 1) if revenue > 0 else 0

                conn = get_db()
                cur  = conn.cursor()
                cur.execute("""
                    INSERT INTO campaign_metrics_history
                        (seller_id, campaign_id, campaign_name, date, spend, revenue, roas, acos, clicks, impressions, orders, cvr)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (seller_id, campaign_id, date) DO UPDATE SET
                        spend=EXCLUDED.spend, revenue=EXCLUDED.revenue,
                        roas=EXCLUDED.roas, acos=EXCLUDED.acos,
                        clicks=EXCLUDED.clicks, impressions=EXCLUDED.impressions,
                        orders=EXCLUDED.orders, cvr=EXCLUDED.cvr,
                        recorded_at=NOW()
                """, (user_id, camp_id, camp_name, yesterday,
                      spend, revenue, roas, acos, clicks, imps, orders, cvr))
                conn.commit()
                cur.close(); conn.close()
                print(f"[HISTORY] {seller.get('nickname',user_id)} / {camp_name} / {yesterday} salvo")
            except Exception as e:
                print(f"[HISTORY][ERRO] campanha {camp_id}:", e)

    print("[HISTORY] Coleta concluída.")


@app.route("/api/ads/<user_id>/campaign/<camp_id>/history")
@login_required
def get_campaign_history(user_id, camp_id):
    ok, err = check_seller_access(user_id)
    if not ok: return err
    days = int(request.args.get("days", 30))
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("""
            SELECT date, spend, revenue, roas, acos, clicks, impressions, orders, cvr
            FROM campaign_metrics_history
            WHERE seller_id=%s AND campaign_id=%s
              AND date >= CURRENT_DATE - INTERVAL '%s days'
            ORDER BY date ASC
        """, (user_id, camp_id, days))
        rows = cur.fetchall()
        cur.close(); conn.close()
        history = [dict(r) for r in rows]
        for h in history:
            h["date"] = h["date"].strftime("%Y-%m-%d") if hasattr(h["date"], "strftime") else str(h["date"])
            for k in ["spend","revenue","roas","acos","cvr"]:
                h[k] = float(h[k]) if h[k] is not None else 0
        return jsonify({"campaign_id": camp_id, "history": history})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/collect-history", methods=["POST"])
@master_required
def trigger_history_collect():
    """Endpoint para disparar coleta manualmente via painel admin."""
    import threading
    threading.Thread(target=collect_campaign_history, daemon=True).start()
    return jsonify({"status": "started"})


# Inicia o scheduler para coleta diária às 06:00
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    scheduler = BackgroundScheduler()
    scheduler.add_job(collect_campaign_history, "cron", hour=6, minute=0, id="daily_history")
    scheduler.start()
    print("[SCHEDULER] Coleta diária de histórico agendada para 06:00")
except Exception as e:
    print("[SCHEDULER][AVISO] APScheduler não disponível:", e)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
